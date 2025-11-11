import io
import csv
import logging
from datetime import datetime
from typing import Optional, List

from openpyxl import load_workbook, Workbook
import xlrd
import xlwt

from models.file_model import FileModel

# Initialize logger
logger = logging.getLogger(__name__)


class ExcelUtilities:
    """
    Utility class for Excel operations
    Supports both XLS and XLSX file formats.
    """
    
    @staticmethod
    def is_xls_file(excel_file: FileModel) -> bool:
        """
        Checks if the file is an XLS file based on extension or magic bytes.
        
        Args:
            excel_file: FileModel containing the Excel file
            
        Returns:
            True if file is XLS format, False otherwise
        """
        # Check by file extension
        filename_lower = excel_file.name.lower()
        if filename_lower.endswith('.xls') and not filename_lower.endswith('.xlsx'):
            return True
        
        # Check by magic bytes (XLS files start with specific bytes)
        if len(excel_file.content) >= 8:
            # XLS files (BIFF format) have specific signatures
            magic_bytes = excel_file.content[:8]
            # Common XLS signatures: D0 CF 11 E0 A1 B1 1A E1 (OLE2 format)
            if magic_bytes[:4] == b'\xd0\xcf\x11\xe0':
                return True
        
        return False

    @staticmethod
    def convert_xls_to_xlsx(excel_file: FileModel, errors: Optional[List[str]] = None) -> Optional[Workbook]:
        """
        Converts an XLS file to an openpyxl Workbook in memory.
        
        Args:
            excel_file: FileModel containing the XLS file
            errors: Optional list to append conversion errors
            
        Returns:
            openpyxl Workbook if conversion successful, None otherwise
        """
        if errors is None:
            errors = []
            
        logger.info(f"Converting XLS file to XLSX format: {excel_file.name}")
        
        try:
            # Load XLS file using xlrd
            xls_book = xlrd.open_workbook(file_contents=excel_file.content)
            sheet_names_list = xls_book.sheet_names() if callable(xls_book.sheet_names) else xls_book.sheet_names
            logger.info(f"Successfully loaded XLS file with {len(sheet_names_list)} sheets: {sheet_names_list}")
            
            # Create new openpyxl workbook
            wb = Workbook()
            # Remove default sheet if we have sheets from XLS
            if len(sheet_names_list) > 0 and wb.active is not None:
                wb.remove(wb.active)
            
            # Convert each sheet
            for sheet_idx, sheet_name in enumerate(sheet_names_list):
                xls_sheet = xls_book.sheet_by_index(sheet_idx)
                
                # Create new sheet in openpyxl workbook
                ws = wb.create_sheet(title=sheet_name)
                
                # Copy all rows
                for row_idx in range(xls_sheet.nrows):
                    for col_idx in range(xls_sheet.ncols):
                        cell_value = xls_sheet.cell_value(row_idx, col_idx)
                        
                        # Handle different cell types from xlrd
                        cell_type = xls_sheet.cell_type(row_idx, col_idx)
                        
                        # Convert xlrd cell types to Python types
                        if cell_type == xlrd.XL_CELL_DATE:
                            # xlrd returns dates as floats, convert to datetime
                            try:
                                if isinstance(cell_value, (int, float)):
                                    date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                                    cell_value = datetime(*date_tuple)
                                else:
                                    # If not numeric, keep as string
                                    cell_value = str(cell_value)
                            except Exception:
                                # If date conversion fails, keep as string
                                cell_value = str(cell_value)
                        elif cell_type == xlrd.XL_CELL_NUMBER:
                            # Keep numbers as-is
                            pass
                        elif cell_type == xlrd.XL_CELL_BOOLEAN:
                            # Keep boolean as-is
                            pass
                        elif cell_type == xlrd.XL_CELL_ERROR:
                            # Convert error to None or empty string
                            cell_value = None
                        elif cell_type == xlrd.XL_CELL_EMPTY:
                            cell_value = None
                        # XL_CELL_TEXT is already a string, no conversion needed
                        
                        # Write to openpyxl sheet (openpyxl uses 1-based indexing)
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                
                logger.debug(f"Converted sheet '{sheet_name}' with {xls_sheet.nrows} rows and {xls_sheet.ncols} columns")
            
            logger.info(f"Successfully converted XLS file to openpyxl Workbook with {len(wb.sheetnames)} sheets")
            return wb
            
        except Exception as e:
            error_msg = f"Error converting XLS file to XLSX format: {str(e)}"
            logger.error(error_msg, exc_info=True)
            if errors is not None:
                errors.append(error_msg)
            return None

    @staticmethod
    def convert_xlsx_to_xls(excel_file: FileModel, errors: Optional[List[str]] = None) -> Optional[FileModel]:
        """
        Converts an XLSX file to XLS format.
        
        Args:
            excel_file: FileModel containing the XLSX file
            errors: Optional list to append conversion errors
            
        Returns:
            FileModel containing the XLS file if conversion successful, None otherwise
        """
        if errors is None:
            errors = []
            
        logger.info(f"Converting XLSX file to XLS format: {excel_file.name}")
        
        try:
            # Load XLSX file using openpyxl
            wb = load_workbook(io.BytesIO(excel_file.content), read_only=True, data_only=True)
            logger.info(f"Successfully loaded XLSX file with {len(wb.sheetnames)} sheets: {wb.sheetnames}")
            
            # Create new xlwt workbook
            xls_book = xlwt.Workbook()
            
            # Convert each sheet
            for sheet_name in wb.sheetnames:
                xlsx_sheet = wb[sheet_name]
                
                # Create new sheet in xlwt workbook
                xls_sheet = xls_book.add_sheet(sheet_name)
                
                # Copy all rows
                for row_idx, row in enumerate(xlsx_sheet.iter_rows(values_only=False)):
                    for col_idx, cell in enumerate(row):
                        cell_value = cell.value
                        
                        # Handle different cell types
                        if cell_value is None:
                            # Empty cell, skip
                            continue
                        elif isinstance(cell_value, datetime):
                            # Convert datetime to xlwt date format
                            # xlwt uses Excel date format (serial date number)
                            try:
                                # Excel date system: days since 1900-01-01 (which is day 1, not day 0)
                                base_date = datetime(1900, 1, 1)
                                delta = cell_value - base_date
                                days_diff = delta.days
                                
                                # Excel incorrectly treats 1900 as a leap year
                                # So dates on or after 1900-03-01 need +1 day adjustment
                                if cell_value >= datetime(1900, 3, 1):
                                    days_diff += 1
                                
                                # Excel counts from 1, not 0 (1900-01-01 is day 1)
                                excel_date = days_diff + 1
                                
                                # Add time component if present (as fraction of day)
                                if delta.seconds > 0 or delta.microseconds > 0:
                                    total_seconds = delta.seconds + (delta.microseconds / 1000000.0)
                                    fraction_of_day = total_seconds / 86400.0
                                    excel_date += fraction_of_day
                                
                                # Create date style (adjust format based on whether time is present)
                                if delta.seconds > 0 or delta.microseconds > 0:
                                    date_style = xlwt.easyxf(num_format_str='YYYY-MM-DD HH:MM:SS')
                                else:
                                    date_style = xlwt.easyxf(num_format_str='YYYY-MM-DD')
                                
                                xls_sheet.write(row_idx, col_idx, excel_date, date_style)
                            except Exception as date_err:
                                # If date conversion fails, write as string
                                logger.warning(f"Date conversion failed for cell ({row_idx}, {col_idx}): {date_err}")
                                xls_sheet.write(row_idx, col_idx, str(cell_value))
                        elif isinstance(cell_value, (int, float)):
                            # Number
                            xls_sheet.write(row_idx, col_idx, cell_value)
                        elif isinstance(cell_value, bool):
                            # Boolean
                            xls_sheet.write(row_idx, col_idx, 1 if cell_value else 0)
                        else:
                            # String or other types
                            xls_sheet.write(row_idx, col_idx, str(cell_value))
                
                logger.debug(f"Converted sheet '{sheet_name}' with {xlsx_sheet.max_row} rows and {xlsx_sheet.max_column} columns")
            
            # Save to bytes
            output = io.BytesIO()
            xls_book.save(output)
            output.seek(0)
            
            # Generate XLS filename
            original_name = excel_file.name
            if original_name.lower().endswith('.xlsx'):
                xls_filename = original_name[:-5] + '.xls'
            else:
                xls_filename = original_name + '.xls'
            
            logger.info(f"Successfully converted XLSX file to XLS format: {xls_filename}")
            return FileModel(name=xls_filename, content=output.getvalue())
            
        except Exception as e:
            error_msg = f"Error converting XLSX file to XLS format: {str(e)}"
            logger.error(error_msg, exc_info=True)
            if errors is not None:
                errors.append(error_msg)
            return None

    @staticmethod
    def load_excel_workbook(excel_file: FileModel, errors: Optional[List[str]] = None, 
                           read_only: bool = False, data_only: bool = True) -> Optional[Workbook]:
        """
        Loads an Excel file (XLS or XLSX) as an openpyxl Workbook.
        Automatically detects and converts XLS files to XLSX format.
        
        Args:
            excel_file: FileModel containing the Excel file
            errors: Optional list to append validation/conversion errors
            read_only: If True, opens workbook in read-only mode (only for XLSX)
            data_only: If True, loads only calculated values (only for XLSX)
            
        Returns:
            openpyxl Workbook if successful, None otherwise
        """
        if errors is None:
            errors = []
            
        logger.info(f"Loading Excel file: {excel_file.name}")
        
        if not excel_file.content or len(excel_file.content) == 0:
            error_msg = "Excel file is empty"
            logger.error(error_msg)
            if errors is not None:
                errors.append(error_msg)
            return None

        # Check if file is XLS format and convert if needed
        if ExcelUtilities.is_xls_file(excel_file):
            logger.info(f"Detected XLS file format, converting to XLSX")
            wb = ExcelUtilities.convert_xls_to_xlsx(excel_file, errors)
            return wb
        else:
            # Try to load as XLSX
            try:
                wb = load_workbook(io.BytesIO(excel_file.content), read_only=read_only, data_only=data_only)
                logger.info(f"Successfully loaded XLSX workbook with {len(wb.sheetnames)} sheets: {wb.sheetnames}")
                return wb
            except Exception as e:
                error_msg = f"Invalid Excel file format: {str(e)}"
                logger.error(error_msg, exc_info=True)
                if errors is not None:
                    errors.append(error_msg)
                return None
    
    @staticmethod
    def excel_to_tsv_files(excel_file: FileModel) -> list[FileModel]:
        """
        Converts an Excel file (XLS or XLSX) to multiple TSV files, one per sheet
        
        Args:
            excel_file: FileModel containing the Excel file
            
        Returns:
            List of FileModels, each containing a TSV file for one sheet
        """
        result = []
        errors = []
        
        # Load the Excel workbook (supports both XLS and XLSX)
        wb = ExcelUtilities.load_excel_workbook(excel_file, errors=errors, read_only=True)
        
        if wb is None:
            logger.error(f"Failed to load Excel file: {excel_file.name}")
            return result
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Create a TSV output buffer
            output = io.StringIO()
            tsv_writer = csv.writer(output, delimiter='\t')
            
            # Write each row to the TSV
            for row in sheet.rows:
                # Extract cell values from the row
                row_values = [cell.value if cell.value is not None else "" for cell in row]
                tsv_writer.writerow(row_values)
            
            # Create a FileModel with the TSV content
            tsv_filename = f"{sheet_name}.tsv"
            tsv_content = output.getvalue().encode('utf-8')
            
            result.append(FileModel(name=tsv_filename, content=tsv_content))
        
        return result