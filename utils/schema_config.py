# Schema Configurations for PyGrays API

from typing import Dict, List, Any, Type, Optional, Union, Callable
import csv
import io
import logging
import re
from datetime import datetime, timedelta
import decimal
from openpyxl import Workbook, worksheet
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

class ImportField:
    def __init__(self, field_type: str, required: bool = False, formats: Optional[List[str]] = None):
        self.field_type = field_type
        self.required = required
        self.formats = formats or []

    def convert(self, value: Any) -> Any:
        if not value:
            return None if not self.required else value
        try:
            if self.field_type == 'datetime':
                return self._parse_date(value)
            elif self.field_type == 'float':
                return float(value)
            elif self.field_type == 'integer':
                return int(value)
            elif self.field_type == 'boolean':
                return value.upper() in ['TRUE', 'YES', 'Y', '1'] if isinstance(value, str) else bool(value)
            elif self.field_type == 'decimal':
                cleaned_value = re.sub(r'[^\d.]', '', str(value))
                return decimal.Decimal(cleaned_value)
            return value
        except (ValueError, TypeError, decimal.InvalidOperation) as e:
            logger.warning(f'Conversion error for value {value} to type {self.field_type}: {str(e)}')
            return value

    def _parse_date(self, date_string: str) -> Optional[datetime]:
        if not date_string:
            return None
        for fmt in self.formats:
            try:
                return datetime.strptime(date_string, fmt)
            except ValueError:
                continue
        logger.warning(f'Failed to parse date {date_string} with formats {self.formats}')
        return None

class ExportField:
    def __init__(self, field_type: str, number_format: Optional[str] = None):
        self.field_type = field_type
        self.number_format = number_format

class ConditionalFormat:
    def __init__(self, column: str, condition: str, reference_value: Any, format_config: Dict[str, Any]):
        self.column = column
        self.condition = condition
        self.reference_value = reference_value
        self.format_config = format_config

    def should_apply(self, cell_value: Any, context: Dict[str, Any]) -> bool:
        """Check if the conditional format should be applied to the cell value"""
        if self.condition == 'date_before_or_equal':
            if isinstance(cell_value, datetime):
                reference_date = context.get(self.reference_value)
                if reference_date:
                    return cell_value.date() <= reference_date
        return False

    def apply_format(self, cell):
        """Apply the formatting to the cell"""
        if 'fill_color' in self.format_config:
            fill_type = self.format_config.get('fill_type', 'solid')
            cell.fill = PatternFill(
                start_color=self.format_config['fill_color'],
                end_color=self.format_config['fill_color'],
                fill_type=fill_type
            )

class BaseSchema:
    def __init__(self, schema: Dict[str, Any]):
        self.schema = schema

class ImportSchema(BaseSchema):
    def import_data(self, raw_data: bytes, errors: List[str]) -> List[Dict[str, Any]]:
        try:
            text = raw_data.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            data = []
            row_count = 0
            conversion_errors = 0

            for row in reader:
                row_count += 1
                converted_row = {}
                row_errors = 0

                for field, value in row.items():
                    if field not in self.schema:
                        converted_row[field] = value
                        continue

                    field_schema = self.schema[field]
                    if not value and field_schema.required:
                        logger.warning(f'Missing required field {field} in row {row_count}')
                        row_errors += 1
                    converted_row[field] = field_schema.convert(value) if value else None

                if row_errors > 0:
                    conversion_errors += 1
                data.append(converted_row)

            logger.info(f'Imported {row_count} rows, with {conversion_errors} conversion errors')
            return data
        except Exception as e:
            errors.append(f'Error importing data: {str(e)}')
            logger.error(f'Error importing data', exc_info=True)
            return []

class ExportSchema(BaseSchema):
    def __init__(self, schema: Dict[str, ExportField], sort_by: Optional[str] = None, 
                 ascending: bool = True, conditional_formats: Optional[List[ConditionalFormat]] = None):
        super().__init__(schema)
        self.sort_by = sort_by
        self.ascending = ascending
        self.conditional_formats = conditional_formats or []

    def export_data(self, data: List[Dict[str, Any]], workbook: Workbook, sheet_name: str, 
                   errors: List[str], context: Optional[Dict[str, Any]] = None) -> bool:
        try:
            sheet = workbook.create_sheet(sheet_name)
            headers = list(self.schema.keys())
            sheet.append(headers)

            # Sort data if sort_by is specified
            sorted_data = data
            if self.sort_by and self.sort_by in headers:
                try:
                    sorted_data = sorted(data, key=lambda row: row.get(self.sort_by, datetime.min) if self.sort_by else datetime.min, reverse=not self.ascending)
                    logger.info(f"Sorted {len(sorted_data)} rows by '{self.sort_by}' ({'ascending' if self.ascending else 'descending'}) for sheet '{sheet_name}'")
                except Exception as sort_error:
                    logger.warning(f"Failed to sort data by '{self.sort_by}' for sheet '{sheet_name}': {str(sort_error)}")
                    sorted_data = data

            for row_idx, item in enumerate(sorted_data, start=2):
                row_values = []
                for col_idx, col in enumerate(headers, start=1):
                    value = item.get(col, '')
                    if isinstance(value, decimal.Decimal):
                        try:
                            value = value.quantize(decimal.Decimal('0.01'), rounding=decimal.ROUND_HALF_UP)
                        except (decimal.InvalidOperation, TypeError):
                            value = ''
                    row_values.append(value)
                
                # First add the row values directly to cells
                for col_idx, value in enumerate(row_values, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    # Apply conditional formatting if specified
                    if self.conditional_formats and context:
                        column_name = headers[col_idx - 1]
                        for conditional_format in self.conditional_formats:
                            if conditional_format.column == column_name:
                                if conditional_format.should_apply(value, context):
                                    conditional_format.apply_format(cell)
                    
                # Then apply number formats if specified
                for col_idx, col in enumerate(headers, start=1):
                    if col in self.schema and self.schema[col].number_format:
                        sheet.cell(row=row_idx, column=col_idx).number_format = self.schema[col].number_format

            logger.info(f'Exported {len(sorted_data)} rows to {sheet_name} sheet')
            if self.conditional_formats:
                logger.info(f'Applied {len(self.conditional_formats)} conditional format(s) to {sheet_name} sheet')
            return True
        except Exception as e:
            errors.append(f'Error exporting data to {sheet_name}: {str(e)}')
            logger.error(f'Error exporting data to {sheet_name}', exc_info=True)
            return False

# Aging Report Service Schemas
aging_report_daily_data_import_schema = ImportSchema({
    'Classification': ImportField('string'),
    'Sale_No': ImportField('string', required=True),
    'Description': ImportField('string'),
    'Division': ImportField('string', required=True),
    'BDM': ImportField('string'),
    'Sale_Date': ImportField('datetime', formats=['%d/%m/%Y %H:%M', '%d/%m/%Y %I:%M:%S %p', '%d/%m/%Y']),
    'Gross_Tot': ImportField('float', required=True),
    'Delot_Ind': ImportField('boolean'),
    'Cheque_Date': ImportField('datetime', formats=['%d/%m/%Y %H:%M', '%d/%m/%Y %I:%M:%S %p', '%d/%m/%Y']),
    'Day0': ImportField('float'),
    'Day1': ImportField('float'),
    'Day2': ImportField('float'),
    'Day3': ImportField('float'),
    'Day4': ImportField('float'),
    'Day5': ImportField('float'),
    'Day6': ImportField('float'),
    'Day7': ImportField('float'),
    'Day8': ImportField('float'),
    'Day9': ImportField('float'),
    'Day10': ImportField('float'),
    'Day11': ImportField('float'),
    'Day12': ImportField('float'),
    'Day13': ImportField('float'),
    'Day14': ImportField('float'),
    'Day15': ImportField('float'),
    'Day16': ImportField('float'),
    'Day17': ImportField('float'),
    'Day18': ImportField('float'),
    'Day19': ImportField('float'),
    'Day20': ImportField('float'),
    'Day21': ImportField('float'),
    'Day22': ImportField('float'),
    'Day23': ImportField('float'),
    'Day24': ImportField('float'),
    'Day25': ImportField('float'),
    'Day26': ImportField('float'),
    'Day27': ImportField('float'),
    'Day28': ImportField('float'),
    'Day29': ImportField('float'),
    'Day30': ImportField('float'),
    'Day31': ImportField('float'),
})

# Base field definitions for aging report exports
aging_report_base_fields = {
    'Classification': ExportField('string'),
    'Sale_No': ExportField('string'),
    'Description': ExportField('string'),
    'Division': ExportField('string'),
    'BDM': ExportField('string'),
    'Sale_Date': ExportField('datetime'),
    'Gross_Tot': ExportField('float'),
    'Delot_Ind': ExportField('boolean'),
    'Cheque_Date': ExportField('datetime'),
    'Day0': ExportField('float'),
    'Day1': ExportField('float'),
    'Day2': ExportField('float'),
    'Day3': ExportField('float'),
    'Day4': ExportField('float'),
    'Day5': ExportField('float'),
    'Day6': ExportField('float'),
    'Day7': ExportField('float'),
    'Day8': ExportField('float'),
    'Day9': ExportField('float'),
    'Day10': ExportField('float'),
    'Day11': ExportField('float'),
    'Day12': ExportField('float'),
    'Day13': ExportField('float'),
    'Day14': ExportField('float'),
    'Day15': ExportField('float'),
    'Day16': ExportField('float'),
    'Day17': ExportField('float'),
    'Day18': ExportField('float'),
    'Day19': ExportField('float'),
    'Day20': ExportField('float'),
    'Day21': ExportField('float'),
    'Day22': ExportField('float'),
    'Day23': ExportField('float'),
    'Day24': ExportField('float'),
    'Day25': ExportField('float'),
    'Day26': ExportField('float'),
    'Day27': ExportField('float'),
    'Day28': ExportField('float'),
    'Day29': ExportField('float'),
    'Day30': ExportField('float'),
    'Day31': ExportField('float'),
    'State': ExportField('string'),
    'State-Division Name': ExportField('string'),
    'Payment Days': ExportField('integer'),
    'Due Date': ExportField('datetime', number_format='DD-MMM-YY'),
    'Division Name': ExportField('string'),
    'Sub Division Name': ExportField('string'),
    'Gross Amount': ExportField('float', number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'),
    'Collected': ExportField('float', number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'),
    'To be Collected': ExportField('float', number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'),
    'Payable to Vendor': ExportField('float', number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'),
    'Month': ExportField('string'),
    'Year': ExportField('integer'),
    'Cheque Date Y/N': ExportField('string'),
    'Days Late for Vendors Pmt': ExportField('integer'),
    'Comments': ExportField('string')
}

# Main data sheet schema (no special formatting)
aging_report_data_schema = ExportSchema(aging_report_base_fields)

# Fully paid sheet schema with sorting and conditional formatting
aging_report_fully_paid_schema = ExportSchema(
    aging_report_base_fields,
    sort_by='Due Date',
    ascending=True,
    conditional_formats=[
        ConditionalFormat(
            column='Due Date',
            condition='date_before_or_equal',
            reference_value='yesterday',
            format_config={'fill_color': '90EE90', 'fill_type': 'solid'}
        )
    ]
)

# Not fully paid sheet schema (no special formatting for now)
aging_report_not_fully_paid_schema = ExportSchema(aging_report_base_fields)

# Legacy schema for backward compatibility
aging_report_export_schema = aging_report_data_schema

# Inventory Service Schemas
inventory_dropship_sales_schema = ImportSchema({
    'Customer': ImportField('string'),
    'AX_ProductCode': ImportField('string'),
    'GST': ImportField('string'),
    'Units': ImportField('integer'),
    'Price': ImportField('decimal'),
    'Amount': ImportField('decimal'),
    'SaleNo': ImportField('string'),
    'VendorNo': ImportField('string'),
    'ItemNo': ImportField('string'),
    'Description': ImportField('string'),
    'Serial_No': ImportField('string'),
    'Vendor_Ref_No': ImportField('string'),
    'DropShipper': ImportField('string'),
    'Consignment': ImportField('string'),
    'DealNo': ImportField('string'),
    'Column1': ImportField('string'),
    'BP': ImportField('decimal'),
    'SaleType': ImportField('string'),
    'FreightCodeDescription': ImportField('string')
})

inventory_deals_schema = ImportSchema({
    'Customer': ImportField('string'),
    'AX_ProductCode': ImportField('string'),
    'GST': ImportField('string'),
    'Units': ImportField('integer'),
    'Price': ImportField('decimal'),
    'Amount': ImportField('decimal'),
    'SaleNo': ImportField('string'),
    'VendorNo': ImportField('string'),
    'ItemNo': ImportField('string'),
    'Description': ImportField('string'),
    'Serial_No': ImportField('string'),
    'Vendor_Ref_No': ImportField('string'),
    'DropShipper': ImportField('string'),
    'Consignment': ImportField('string'),
    'DealNo': ImportField('string'),
    'Column1': ImportField('string'),
    'BP': ImportField('decimal'),
    'SaleType': ImportField('string'),
    'DivisionCode': ImportField('string'),
    'DivisionDescription': ImportField('string'),
    'FreightCodeDescription': ImportField('string')
})

inventory_uom_mapping_schema = ImportSchema({
    'Item': ImportField('string'),
    'UOM': ImportField('decimal')
})

inventory_mixed_export_schema = ExportSchema({
    'Customer': ExportField('string'),
    'AX_ProductCode': ExportField('string'),
    'Per_Unit_Cost': ExportField('decimal'),
    'Units': ExportField('integer'),
    'Price': ExportField('decimal'),
    'Amount': ExportField('decimal'),
    'SaleNo': ExportField('string'),
    'VendorNo': ExportField('string'),
    'ItemNo': ExportField('string'),
    'Description': ExportField('string'),
    'Serial_No': ExportField('string'),
    'COGS': ExportField('decimal'),
    'SALE_EX_GST': ExportField('decimal'),
    'BP_EX_GST': ExportField('decimal'),
    'Vendor_Ref_No': ExportField('string'),
    'DropShipper': ExportField('string'),
    'Consignment': ExportField('string'),
    'DealNo': ExportField('string'),
    'Column1': ExportField('string'),
    'BP': ExportField('decimal'),
    'SaleType': ExportField('string'),
    'FreightCodeDescription': ExportField('string')
})

inventory_wine_export_schema = ExportSchema({
    'Customer': ExportField('string'),
    'AX_ProductCode': ExportField('string'),
    'Per_Unit_Cost': ExportField('decimal'),
    'Units': ExportField('integer'),
    'Price': ExportField('decimal'),
    'Amount': ExportField('decimal'),
    'SaleNo': ExportField('string'),
    'VendorNo': ExportField('string'),
    'ItemNo': ExportField('string'),
    'Description': ExportField('string'),
    'Serial_No': ExportField('string'),
    'COGS': ExportField('decimal'),
    'SALE_EX_GST': ExportField('decimal'),
    'BP_EX_GST': ExportField('decimal'),
    'Vendor_Ref_No': ExportField('string'),
    'DropShipper': ExportField('string'),
    'Consignment': ExportField('string'),
    'DealNo': ExportField('string'),
    'Column1': ExportField('string'),
    'BP': ExportField('decimal'),
    'SaleType': ExportField('string'),
    'DivisionCode': ExportField('string'),
    'DivisionDescription': ExportField('string'),
    'FreightCodeDescription': ExportField('string')
}) 