import csv
import io
import logging
import re
from typing import List, Tuple, Dict, Optional
from datetime import datetime

from openpyxl import Workbook
from models.response_base import ResponseBase
from models.file_model import FileModel
from utils.schema_config import (
    inventory_dropship_sales_schema as inventory_dropship_sales_import_schema,
    inventory_deals_schema as inventory_deals_import_schema,
    inventory_uom_mapping_schema as inventory_uom_mapping_import_schema,
    inventory_dropship_sales_schema as inventory_dropship_sales_export_schema,
    inventory_mixed_export_schema,
    inventory_wine_export_schema, BaseSchema
)

import decimal

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class InventoryService:
    # Use schemas from configuration
    dropship_sales_import_schema = inventory_dropship_sales_import_schema
    deals_import_schema = inventory_deals_import_schema
    uom_mapping_import_schema:BaseSchema = inventory_uom_mapping_import_schema
    dropship_sales_export_schema = inventory_dropship_sales_export_schema
    mixed_export_schema = inventory_mixed_export_schema
    wine_export_schema = inventory_wine_export_schema

    def __init__(self):
        pass

    @staticmethod
    def _load_csv_from_bytes(csv_data: bytes) -> list[list[str]]:
        bio = io.BytesIO(csv_data)
        
        # Try different encodings with BOM handling
        encodings_to_try = [
            ('utf-8-sig', 'UTF-8 with BOM handling'),
            ('utf-8', 'standard UTF-8'),
            ('latin-1', 'fallback encoding')
        ]
        
        last_error = None
        for encoding, desc in encodings_to_try:
            bio.seek(0)
            try:
                logger.debug(f"Trying to decode with {desc}")
                text_wrapper = io.TextIOWrapper(bio, encoding=encoding, newline='')
                sample = text_wrapper.read(4096)
                if not sample:
                    logger.debug(f"Empty sample with {desc}, trying next encoding")
                    continue
                    
                text_wrapper.seek(0)
                
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',\t')
                    logger.debug(f"Successfully sniffed CSV dialect with {desc}")
                except csv.Error:
                    logger.debug(f"Could not determine dialect with {desc}, using excel dialect")
                    dialect = csv.excel
    
                reader = csv.reader(text_wrapper, dialect=dialect)
                rows = list(reader)
                if rows:
                    logger.debug(f"Successfully read {len(rows)} rows with {desc}")
                    return rows
                logger.debug(f"No rows read with {desc}, trying next encoding")
            except Exception as e:
                last_error = e
                logger.debug(f"Error with {desc}: {str(e)}")
                continue
        
        # If we get here, all encoding attempts failed
        error_msg = f"Failed to decode CSV data with any encoding: {str(last_error)}"
        logger.error(error_msg)
        raise ValueError(error_msg)

    @staticmethod
    def _extract_and_validate_file_dates(file_names: List[str]) -> Tuple[Optional[int], Optional[int], bool]:
        """
        Extract month and year from file names and validate they are consistent.
        
        Args:
            file_names: List of file names to process.
            
        Returns:
            Tuple containing:
            - Extracted month (or None if invalid)
            - Extracted year (or None if invalid)
            - Boolean indicating if validation was successful
        """
        consensus_month = None
        consensus_year = None
        
        if not file_names:
            return None, None, False
    
        for file_name in file_names:
            # decompose file name
            # E.g., DropshipSales20250228.txt will be decomposed into Category: DropshipSales, Date: 20250228, Extension: txt
    
            pattern = r'^(.+)(\d{8})\.([a-zA-Z]+)$'
            match = re.match(pattern, file_name)
            if not match:
                return None, None, False
    
            category = match.group(1)  # Group 1: Category
            date = match.group(2)  # Group 2: Date
            extension = match.group(3)  # Group 3: Extension
    
            # Get month and year from date
            month = int(date[4:6])
            year = int(date[0:4])
            
            # Validate month range
            if month < 1 or month > 12:
                return None, None, False
    
            # Ensure that all files have the same month and year
            if consensus_month is None:
                consensus_month = month
            else:
                if month != consensus_month:
                    return consensus_month, consensus_year, False
    
            if consensus_year is None:
                consensus_year = year
            else:
                if year != consensus_year:
                    return consensus_month, consensus_year, False
    
        return consensus_month, consensus_year, True

    @staticmethod
    def _extract_date_from_soh_filename(filename: str) -> Optional[datetime]:
        """
        Extract date from SOH filename ending with DDMMYY pattern.
        
        Args:
            filename: Filename (with or without extension)
            
        Returns:
            datetime object if valid date found, None otherwise
        """
        # Remove extension if present
        name_without_ext = filename.rsplit('.', 1)[0] if '.' in filename else filename
        
        # Check if filename ends with 6 digits (DDMMYY)
        if len(name_without_ext) < 6 or not name_without_ext[-6:].isdigit():
            return None
            
        date_str = name_without_ext[-6:]
        
        try:
            # Parse DDMMYY format
            day = int(date_str[0:2])
            month = int(date_str[2:4])
            year = int(date_str[4:6])
            
            # Convert 2-digit year to 4-digit year
            # Assume years 00-30 are 2000-2030, years 31-99 are 1931-1999
            if year <= 30:
                year += 2000
            else:
                year += 1900
                
            # Validate and create datetime
            parsed_date = datetime(year, month, day)
            return parsed_date
            
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _validate_soh_filename(filename: str) -> bool:
        """
        Validate that filename ends with DDMMYY pattern and has valid date.
        
        Args:
            filename: Filename to validate
            
        Returns:
            True if filename is valid SOH file, False otherwise
        """
        return InventoryService._extract_date_from_soh_filename(filename) is not None
        
    def _get_mixed_deals(self, data_dicts: List[Dict]) -> List[Dict]:
        """
        Extract records with MIXED deal numbers from the dataset.
        
        Args:
            data_dicts: List of data dictionaries to filter.
            
        Returns:
            List of dictionaries with MIXED deal numbers.
        """
        mixed_deals = [item for item in data_dicts if str(item.get("Customer", "")) == "10" and item.get("AX_ProductCode", "").strip() != ""]
        logger.info(f"Found {len(mixed_deals)} rows with MIXED DealNo")
        return mixed_deals
        
    def _add_per_unit_cost(self, data_dicts: List[Dict], soh_files_data: List[Dict], errors: List[str]) -> List[Dict]:
        """
        Add Per_Unit_Cost field to each item using fallback lookup through SOH files.
        
        Args:
            data_dicts: List of data dictionaries to enrich.
            soh_files_data: List of SOH file data sorted by date (newest first).
            errors: List to collect error messages.
            
        Returns:
            List of dictionaries with Per_Unit_Cost added.
        """
        items_not_found = []
        lookup_stats = {}
        
        for item in data_dicts:
            product_code = item.get("AX_ProductCode", "")
            per_unit_cost = ""
            found_in_file = None
            
            # Try to find the item in SOH files (newest first)
            for file_data in soh_files_data:
                if product_code in file_data['mapping']:
                    per_unit_cost = file_data['mapping'][product_code]
                    found_in_file = file_data['filename']
                    break
            
            if found_in_file:
                # Track which file provided the lookup
                if found_in_file not in lookup_stats:
                    lookup_stats[found_in_file] = 0
                lookup_stats[found_in_file] += 1
            else:
                # Item not found in any SOH file
                items_not_found.append(product_code)
            
            item["Per_Unit_Cost"] = per_unit_cost
        
        # Log lookup statistics
        logger.info(f"Added Per_Unit_Cost field to {len(data_dicts)} items")
        for filename, count in lookup_stats.items():
            logger.info(f"  {count} items found in {filename}")
        
        # Handle items not found in any SOH file
        if items_not_found:
            error_msg = f"Items not found in any SOH file: {', '.join(items_not_found[:10])}"
            if len(items_not_found) > 10:
                error_msg += f" (and {len(items_not_found) - 10} more)"
            logger.error(error_msg)
            errors.append(error_msg)
            
            # Log detailed information about missing items
            logger.error(f"Total items not found: {len(items_not_found)}")
            logger.error(f"Available SOH files: {[f['filename'] for f in soh_files_data]}")
        
        return data_dicts
    
    def _calculate_additional_fields(self, data_dicts: List[Dict]) -> List[Dict]:
        """
        Calculate and add COGS, SALE_EX_GST, and BP_EX_GST fields to data dictionaries.
        
        Args:
            data_dicts: List of data dictionaries to enrich.
            
        Returns:
            List of dictionaries with calculated fields added.
        """
        for item in data_dicts:
            # Calculate COGS = Per_Unit_Cost * Units
            per_unit_cost = item.get("Per_Unit_Cost", "")
            units_value = item.get("Units", 0)
            
            if per_unit_cost and units_value:
                try:
                    per_unit_cost_decimal = decimal.Decimal(per_unit_cost) if per_unit_cost else decimal.Decimal('0')
                    cogs_decimal = per_unit_cost_decimal * decimal.Decimal(units_value)
                    item["COGS"] = cogs_decimal
                except (decimal.InvalidOperation, TypeError):
                    item["COGS"] = ""
            else:
                item["COGS"] = ""
            
            # Calculate SALE_EX_GST = Amount / 1.1
            amount_value = item.get("Amount", "")
            if amount_value and isinstance(amount_value, decimal.Decimal):
                try:
                    sale_ex_gst_decimal = amount_value / decimal.Decimal('1.1')
                    item["SALE_EX_GST"] = sale_ex_gst_decimal
                except (decimal.InvalidOperation, TypeError):
                    item["SALE_EX_GST"] = ""
            else:
                item["SALE_EX_GST"] = ""
            
            # Calculate BP_EX_GST = BP / 1.1
            bp_value = item.get("BP", "")
            if bp_value and isinstance(bp_value, decimal.Decimal):
                try:
                    bp_ex_gst_decimal = bp_value / decimal.Decimal('1.1')
                    item["BP_EX_GST"] = bp_ex_gst_decimal
                except (decimal.InvalidOperation, TypeError):
                    item["BP_EX_GST"] = ""
            else:
                item["BP_EX_GST"] = ""
                
        logger.info(f"Calculated additional fields for {len(data_dicts)} items")
        return data_dicts
        
    def _handle_errors(self, errors: List[str], response: ResponseBase) -> bool:
        """
        Helper method to check for errors and update response accordingly.
        
        Args:
            errors: List of error messages to check.
            response: Response object to update if errors exist.
            
        Returns:
            True if errors exist, False otherwise.
        """
        if len(errors) > 0:
            response.errors.extend(errors)
            response.is_success = False
            return True
        return False
        
    def _prepare_workbook(self, errors: List[str]) -> Optional[Workbook]:
        """
        Create and prepare a new Excel workbook with necessary sheets.
        
        Args:
            errors: List to collect error messages.
            
        Returns:
            Prepared Workbook object or None if error occurs.
        """
        try:
            new_workbook = Workbook()
            
            # Create sheets for data
            dropship_sales_sheet = new_workbook.create_sheet("Dropship Sales")
            mixed_sheet = new_workbook.create_sheet("Mixed")
            wine_sheet = new_workbook.create_sheet("Wine")
            
            # Remove the default sheet if it exists
            if 'Sheet' in new_workbook.sheetnames:
                del new_workbook['Sheet']
            
            return new_workbook
        except Exception as e:
            errors.append(f"Error creating workbook: {str(e)}")
            logger.error("Error creating workbook", exc_info=True)
            return None
            
    def _write_dropship_sales_sheet(self, workbook: Workbook, data_dicts: List[Dict], errors: List[str]) -> bool:
        """
        Write dropship sales data to the Excel sheet using export schema.
        
        Args:
            workbook: The workbook to write to.
            data_dicts: Data to write to the sheet.
            errors: List to collect error messages.
            
        Returns:
            True if successful, False if errors occurred.
        """
        try:
            sheet = workbook['Dropship Sales']
            headers = list(self.dropship_sales_export_schema.schema.keys())
            sheet.append(headers)
            for item in data_dicts:
                row_values = []
                for col in headers:
                    value = item.get(col, '')
                    if isinstance(value, decimal.Decimal):
                        try:
                            value = value.quantize(decimal.Decimal('0.01'), rounding=decimal.ROUND_HALF_UP)
                        except (decimal.InvalidOperation, TypeError):
                            value = ''
                    row_values.append(value)
                sheet.append(row_values)
            logger.info(f'Wrote {len(data_dicts)} rows to Dropship Sales sheet')
            return True
        except Exception as e:
            errors.append(f'Error writing to Dropship Sales sheet: {str(e)}')
            logger.error('Error writing to Dropship Sales sheet', exc_info=True)
            return False
            
    def _write_mixed_sheet(self, workbook: Workbook, mixed_deals: List[Dict], errors: List[str]) -> bool:
        """
        Write mixed deals data to the Mixed sheet using pre-calculated fields.
        
        Args:
            workbook: The workbook to write to.
            mixed_deals: Mixed deals data to write.
            errors: List to collect error messages.
            
        Returns:
            True if successful, False if errors occurred.
        """
        try:
            sheet = workbook['Mixed']
            headers = list(self.mixed_export_schema.schema.keys())
            sheet.append(headers)
            for item in mixed_deals:
                row_values = []
                for col in headers:
                    value = item.get(col, '')
                    if isinstance(value, decimal.Decimal):
                        try:
                            value = value.quantize(decimal.Decimal('0.01'), rounding=decimal.ROUND_HALF_UP)
                        except (decimal.InvalidOperation, TypeError):
                            value = ''
                    row_values.append(value)
                sheet.append(row_values)
            logger.info(f'Wrote {len(mixed_deals)} rows to Mixed sheet')
            return True
        except Exception as e:
            errors.append(f'Error writing to Mixed sheet: {str(e)}')
            logger.error('Error writing to Mixed sheet', exc_info=True)
            return False
    
    def _write_wine_sheet(self, workbook: Workbook, data_dicts: List[Dict], errors: List[str]) -> bool:
        """
        Write wine deals data to the Wine sheet using pre-calculated fields.
        
        Args:
            workbook: The workbook to write to.
            data_dicts: Data to write to the sheet.
            errors: List to collect error messages.
            
        Returns:
            True if successful, False if errors occurred.
        """
        try:
            sheet = workbook['Wine']
            headers = list(self.wine_export_schema.schema.keys())
            sheet.append(headers)
            for item in data_dicts:
                row_values = []
                for col in headers:
                    value = item.get(col, '')
                    if isinstance(value, decimal.Decimal):
                        try:
                            value = value.quantize(decimal.Decimal('0.01'), rounding=decimal.ROUND_HALF_UP)
                        except (decimal.InvalidOperation, TypeError):
                            value = ''
                    row_values.append(value)
                sheet.append(row_values)
            logger.info(f'Wrote {len(data_dicts)} rows to Wine sheet')
            return True
        except Exception as e:
            errors.append(f'Error writing to Wine sheet: {str(e)}')
            logger.error('Error writing to Wine sheet', exc_info=True)
            return False

    def _load_csv_data(self, file: FileModel, schema :BaseSchema , errors: List[str]) -> List[Dict]:
        """
        Load and validate CSV data based on a given schema.
    
        Args:
            file: The CSV file to process.
            schema: Dictionary mapping column names to expected data types.
            errors: List to collect error messages.
    
        Returns:
            List of validated row dictionaries. Empty if validation fails.
        """
        try:
            rows = self._load_csv_from_bytes(file.content)
            if not rows:
                errors.append(f'No data in file {file.name}')
                return []

            headers = rows[0]
            schema_dict = schema.schema if hasattr(schema, 'schema') else {}
            missing_columns = [col for col in schema_dict.keys() if col not in headers]
            if missing_columns:
                error_message = f'Missing required columns in {file.name}: {", ".join(missing_columns)}'
                errors.append(error_message)
                return []

            validated_rows = []

            for row_index, row in enumerate(rows[1:], start=2):  # Skip header
                if len(row) != len(headers):
                    errors.append(f'Row {row_index} in {file.name}: mismatched number of columns')
                    continue

                item = {}
                row_invalid = False
                for i, header_name in enumerate(headers):
                    value = row[i] if i < len(row) else ''
                    if header_name in schema_dict:
                        expected_type = schema_dict[header_name].field_type if hasattr(schema_dict[header_name], 'field_type') else None
                        if value and expected_type:
                            if expected_type == 'decimal':
                                try:
                                    # remove anything but digits and decimal point from the string
                                    value = re.sub(r'[^\d.]', '', value)
                                    value = decimal.Decimal(value)
                                except decimal.InvalidOperation:
                                    errors.append(f'Row {row_index}, column \'{header_name}\' in {file.name}: invalid decimal value \'{value}\'')
                                    row_invalid = True
                                    break  # Skip this row
                            elif expected_type == 'integer':
                                try:
                                    # remove anything but digits from the string
                                    value = re.sub(r'\D', '', value)
                                    value = int(value) if value else 0
                                except ValueError:
                                    errors.append(f'Row {row_index}, column \'{header_name}\' in {file.name}: invalid integer value \'{value}\'')
                                    row_invalid = True
                                    break  # Skip this row
                    item[header_name] = value
                if not row_invalid:
                    validated_rows.append(item)

            return validated_rows
        except Exception as e:
            errors.append(f'Error processing file {file.name}: {str(e)}')
            logger.error(f'Error processing file {file.name}', exc_info=True)
            return []

    def _load_multiple_soh_files(self, csv_files: List[FileModel], errors: List[str]) -> Optional[List[Dict[str, str]]]:
        """
        Load and sort multiple SOH files by date (newest first).
        
        Args:
            csv_files: List of CSV files containing SOH data.
            errors: List to collect error messages.
            
        Returns:
            List of dictionaries (one per file) sorted by date (newest first), or None if validation fails.
        """
        logger.info(f"Loading {len(csv_files)} SOH files")
        
        # Validate all filenames first
        invalid_files = []
        for csv_file in csv_files:
            if not self._validate_soh_filename(csv_file.name):
                invalid_files.append(csv_file.name)
        
        if invalid_files:
            error_msg = f"Invalid SOH filename(s) - must end with DDMMYY pattern: {', '.join(invalid_files)}"
            logger.error(error_msg)
            errors.append(error_msg)
            return None
        
        # Load and parse each file
        soh_files_data = []
        for csv_file in csv_files:
            logger.info(f"Loading SOH file: {csv_file.name}")
            
            # Load CSV data
            rows = self._load_csv_data(csv_file, self.uom_mapping_import_schema, errors)
            if errors:
                logger.error(f"Errors loading SOH file {csv_file.name}: {errors}")
                return None
            
            # Build item mapping for this file
            item_uom_map = {}
            conflicting_items = {}
            
            for row in rows:
                item_number = row["Item"]
                uom = row["UOM"]
                if item_number in item_uom_map:
                    if item_uom_map[item_number] != uom:
                        if item_number not in conflicting_items:
                            conflicting_items[item_number] = [item_uom_map[item_number]]
                        conflicting_items[item_number].append(uom)
                else:
                    item_uom_map[item_number] = uom
            
            # Check for conflicts within this file
            if conflicting_items:
                conflict_errors = [f"Item {item} has conflicting UOM values in {csv_file.name}: {', '.join(map(str, uoms))}"
                                 for item, uoms in conflicting_items.items()]
                error_message = f"SOH file {csv_file.name} contains duplicate item numbers with different UOM values"
                logger.error(error_message)
                for err in conflict_errors:
                    logger.error(err)
                errors.append(error_message)
                errors.extend(conflict_errors)
                return None
            
            # Extract date and store file data
            file_date = self._extract_date_from_soh_filename(csv_file.name)
            soh_files_data.append({
                'filename': csv_file.name,
                'date': file_date,
                'mapping': item_uom_map
            })
            
            logger.info(f"Loaded {len(item_uom_map)} items from {csv_file.name} (date: {file_date.strftime('%Y-%m-%d')})")
        
        # Sort by date (newest first)
        soh_files_data.sort(key=lambda x: x['date'], reverse=True)
        
        logger.info(f"Sorted {len(soh_files_data)} SOH files by date (newest first)")
        for file_data in soh_files_data:
            logger.info(f"  {file_data['filename']} - {file_data['date'].strftime('%Y-%m-%d')}")
        
        return soh_files_data

    def _process_deals_files(self, txt_files: List[FileModel], errors: List[str]) -> Tuple[Optional[List[Dict]], Optional[int], Optional[int]]:
        """
        Process Deals files to extract and validate data.
    
        Args:
            txt_files: List of text files to process.
            errors: List to collect error messages.
    
        Returns:
            Tuple containing:
            - List of dictionaries with validated data, or None if validation fails
            - Month extracted from file names, or None if validation fails
            - Year extracted from file names, or None if validation fails
        """
        logger.info("Processing Deals files")
    
        deals_files = sorted([x for x in txt_files if re.match(r'^Deals\d{8}\.txt$', x.name)],
                              key=lambda x: x.name)
        
        if not deals_files:
            errors.append("No Deals files found in the provided files")
            return None, None, None
        
        # Extract and validate the month and year from file names
        month, year, is_valid = self._extract_and_validate_file_dates([x.name for x in deals_files])
        if not is_valid:
            errors.append("Invalid file names - all Deals files must have the same month and year")
            return None, None, None
    
        all_items = []
        for file in deals_files:
            items = self._load_csv_data(file, self.deals_import_schema, errors)
            if len(errors) > 0:
                logger.error(f"Errors processing {file.name}: {errors}")
                return None, None, None
            all_items.extend(items)
    
        logger.info(f"Processed {len(all_items)} rows of data from {len(deals_files)} files")
        return all_items, month, year
    
    def _process_dropship_sales_files(self, txt_files: List[FileModel], errors: List[str]) -> Tuple[Optional[List[Dict]], Optional[int], Optional[int]]:
        """
        Process DropshipSales files to extract and validate data.
    
        Args:
            txt_files: List of text files to process.
            errors: List to collect error messages.
    
        Returns:
            Tuple containing:
            - List of dictionaries with validated data, or None if validation fails
            - Month extracted from file names, or None if validation fails
            - Year extracted from file names, or None if validation fails
        """
        logger.info("Processing DropshipSales files")
    
        dropship_sales_files = sorted([x for x in txt_files if re.match(r'^DropshipSales\d{8}\.txt$', x.name)],
                                      key=lambda x: x.name)
        
        if not dropship_sales_files:
            errors.append("No DropshipSales files found in the provided files")
            return None, None, None
        
        # Extract and validate the month and year from file names
        month, year, is_valid = self._extract_and_validate_file_dates([x.name for x in dropship_sales_files])
        if not is_valid:
            errors.append("Invalid file names - all DropshipSales files must have the same month and year")
            return None, None, None
    
        all_items = []
        for file in dropship_sales_files:
            items = self._load_csv_data(file, self.dropship_sales_import_schema, errors)
            if len(errors) > 0:
                logger.error(f"Errors processing {file.name}: {errors}")
                return None, None, None
            all_items.extend(items)
    
        logger.info(f"Processed {len(all_items)} rows of data from {len(dropship_sales_files)} files")
        return all_items, month, year

    def _get_month_name(self, month: int) -> str:
        """
        Convert month number to month name.
        
        Args:
            month: Month number (1-12)
            
        Returns:
            Month name (e.g., "January", "February", etc.)
        """
        month_names = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]
        if 1 <= month <= 12:
            return month_names[month - 1]
        return "Unknown"
    
    def process_inventory_request(self, txt_files: List[FileModel], csv_files: List[FileModel]) -> ResponseBase:
        """
        Process inventory request by handling input files, generating data, and creating an Excel file.
        
        Args:
            txt_files: List of text files containing dropship sales and deals data.
            csv_files: List of CSV files containing SOH data (must end with DDMMYY pattern).
            
        Returns:
            Response object with success status, Excel file, and any errors.
        """
        logger.info("Processing inventory request")
        response = ResponseBase()
        errors = []  # For collecting errors from all operations
        
        # =====================================================================
        # STAGE 1: Process input files, validate columns and data
        # =====================================================================
        logger.info("Stage 1: Processing and validating input files")
        
        # Validate that we have at least one SOH file
        if not csv_files:
            error_msg = "At least one SOH file (CSV) is required"
            errors.append(error_msg)
            logger.error(error_msg)
            self._handle_errors(errors, response)
            return response
        
        # Load and validate multiple SOH files
        soh_files_data = self._load_multiple_soh_files(csv_files, errors)
        if self._handle_errors(errors, response):
            return response
        
        # Process DropshipSales files
        dropship_sales_data, dropship_month, dropship_year = self._process_dropship_sales_files(txt_files, errors)
        if self._handle_errors(errors, response):
            return response
            
        # Process Deals files
        deals_data, deals_month, deals_year = self._process_deals_files(txt_files, errors)
        if self._handle_errors(errors, response):
            return response
        
        # Validate both types of files have the same month and year
        if dropship_month is not None and deals_month is not None:
            if dropship_month != deals_month or dropship_year != deals_year:
                error_msg = f"Files from different periods: DropshipSales files are from {self._get_month_name(dropship_month)} 20{dropship_year}, while Deals files are from {self._get_month_name(deals_month)} 20{deals_year}"
                errors.append(error_msg)
                logger.error(error_msg)
                self._handle_errors(errors, response)
                return response
        
        # Get month and year for file naming
        file_month = dropship_month or deals_month
        file_year = dropship_year or deals_year
        
        if file_month is None or file_year is None:
            error_msg = "Could not determine month and year from file names"
            errors.append(error_msg)
            logger.error(error_msg)
            self._handle_errors(errors, response)
            return response
            
        # =====================================================================
        # STAGE 2: Generate/calculate extra data
        # =====================================================================
        logger.info("Stage 2: Generating and calculating extra data")
        
        # Extract mixed deals for separate processing
        mixed_deals = self._get_mixed_deals(dropship_sales_data or [])
    
        # Add Per_Unit_Cost to all data sets using fallback lookup
        if mixed_deals:
            self._add_per_unit_cost(mixed_deals, soh_files_data, errors)
            if self._handle_errors(errors, response):
                return response
                
        if deals_data:
            self._add_per_unit_cost(deals_data, soh_files_data, errors)
            if self._handle_errors(errors, response):
                return response
            
        # Calculate additional fields (COGS, SALE_EX_GST, BP_EX_GST)
        if mixed_deals:
            self._calculate_additional_fields(mixed_deals)
        if deals_data:
            self._calculate_additional_fields(deals_data)
    
        # =====================================================================
        # STAGE 3: Prepare Excel sheets and write data
        # =====================================================================
        logger.info("Stage 3: Preparing Excel file and writing data")
        
        # Prepare workbook with necessary sheets
        new_workbook_opt = self._prepare_workbook(errors)
        if not new_workbook_opt:
            if self._handle_errors(errors, response):
                return response
        new_workbook = new_workbook_opt
            
        # Write data to Dropship Sales sheet
        if dropship_sales_data and new_workbook:
            if not self._write_dropship_sales_sheet(new_workbook, dropship_sales_data, errors):
                if self._handle_errors(errors, response):
                    return response
            
        # Write mixed deals data to Mixed sheet
        if mixed_deals and new_workbook:
            if not self._write_mixed_sheet(new_workbook, mixed_deals, errors):
                if self._handle_errors(errors, response):
                    return response
            
        # Write deals data to Wine sheet
        if deals_data and new_workbook:
            if not self._write_wine_sheet(new_workbook, deals_data, errors):
                if self._handle_errors(errors, response):
                    return response
            
        # Save the workbook and prepare response
        try:
            # Create a descriptive file name
            month_name = self._get_month_name(file_month)
            file_name = f'{month_name}_All_Sales_{file_year}.xlsx'
            workbook_bytes = io.BytesIO()
            if new_workbook:
                new_workbook.save(workbook_bytes)
                workbook_binary = workbook_bytes.getvalue()
                response.data = FileModel(name=file_name, content=workbook_binary)
                logger.info(f'Excel workbook saved as {file_name} with Dropship Sales, Mixed, and Wine sheets')
            return response
        except Exception as e:
            error_msg = f'Error saving workbook: {str(e)}'
            logger.error(error_msg, exc_info=True)
            errors.append(error_msg)
            self._handle_errors(errors, response)
            return response