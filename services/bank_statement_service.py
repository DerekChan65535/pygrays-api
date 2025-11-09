import io
import itertools
import csv
import zipfile
from typing import List, Dict, Any, Optional
from decimal import Decimal

from openpyxl import Workbook

from models.file_model import FileModel
from models.response_base import ResponseBase
from services.multi_logging import LoggingService

# Initialize logger
logger = LoggingService().get_logger(__name__)


class BankStatementService:
    """
    Service for processing Westpac Bank Statement CSV files.
    Filters by account numbers, groups by date, and generates Excel files.
    """

    # Hardcoded configuration
    REQUIRED_ACCOUNTS = [
        "032075843041",
        "030162001011700001",
        "034003431178",
        "034008460699",
        "032075842049",
        "034702307846",
        "032075840422",
        "036011606934"
    ]
    
    REQUIRED_CSV_COLUMNS = [
        "TRAN_DATE",
        "ACCOUNT_NO",
        "ACCOUNT_NAME",
        "CCY",
        "CLOSING_BAL",
        "AMOUNT",
        "TRAN_CODE",
        "NARRATIVE",
        "SERIAL"
    ]
    
    NUMERIC_COLUMNS = [
        "TRAN_DATE",
        "AMOUNT",
        "CLOSING_BAL"
    ]

    def __init__(self):
        pass

    def _handle_errors(self, errors: List[str], response: ResponseBase) -> bool:
        """
        Helper method to check for errors and update the response accordingly.
        
        Args:
            errors: List of error messages to check.
            response: Response object to update if errors exist.
            
        Returns:
            True if errors exist, False otherwise.
        """
        if len(errors) > 0:
            logger.warning(f"Handling {len(errors)} errors in response")
            for idx, err in enumerate(errors):
                logger.error(f"Error {idx+1}: {err}")
            response.errors.extend(errors)
            response.is_success = False
            return True
        logger.debug("No errors to handle in response")
        return False

    def _validate_csv_file(self, csv_file: FileModel, errors: List[str]) -> bool:
        """
        Validates that the CSV file is not empty.
        
        Args:
            csv_file: FileModel containing the CSV file
            errors: List to append validation errors
            
        Returns:
            True if valid, False otherwise
        """
        logger.info(f"Validating CSV file: {csv_file.name}")
        
        if not csv_file.content or len(csv_file.content) == 0:
            error_msg = "CSV file is empty"
            logger.error(error_msg)
            errors.append(error_msg)
            return False
        
        logger.info("CSV file validation passed")
        return True

    def _load_csv_data(self, csv_file: FileModel, errors: List[str]) -> Optional[List[Dict[str, Any]]]:
        """
        Loads and parses CSV data into a list of dictionaries.
        Validates headers match expected columns.
        
        Args:
            csv_file: FileModel containing the CSV file
            errors: List to append validation errors
            
        Returns:
            List of dictionaries representing rows, or None if validation fails
        """
        logger.info(f"Loading CSV data from {csv_file.name}")
        
        try:
            # Decode CSV content (handle BOM if present)
            text = csv_file.content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            
            # Get headers from the CSV
            headers = reader.fieldnames
            if headers is None:
                error_msg = "CSV file has no headers"
                logger.error(error_msg)
                errors.append(error_msg)
                return None
            
            # Clean headers (strip whitespace)
            headers = [h.strip() if h else "" for h in headers]
            logger.info(f"Read headers: {headers}")
            logger.info(f"Expected headers: {self.REQUIRED_CSV_COLUMNS}")
            
            # Validate headers match expected columns
            if headers != self.REQUIRED_CSV_COLUMNS:
                error_msg = f"Invalid CSV headers. Expected {self.REQUIRED_CSV_COLUMNS}, got {headers}"
                logger.error(error_msg)
                errors.append(error_msg)
                return None
            
            # Read all rows
            data = []
            row_count = 0
            for row in reader:
                row_count += 1
                # Create dictionary with cleaned values
                row_dict = {col: row.get(col, "").strip() if row.get(col) else "" for col in headers}
                data.append(row_dict)
            
            logger.info(f"Data loaded successfully. Total rows: {row_count}")
            return data
            
        except Exception as e:
            error_msg = f"Error loading CSV data: {str(e)}"
            logger.error(error_msg, exc_info=True)
            errors.append(error_msg)
            return None

    def _filter_by_accounts(self, data: List[Dict[str, Any]], errors: List[str]) -> List[Dict[str, Any]]:
        """
        Filters transactions for required accounts.
        
        Args:
            data: List of transaction dictionaries
            errors: List to append errors (currently unused but kept for consistency)
            
        Returns:
            Filtered list of transactions
        """
        logger.info(f"Filtering data for required accounts: {self.REQUIRED_ACCOUNTS}")
        logger.info(f"Original data size: {len(data)}")
        
        filtered_data = [row for row in data if row.get('ACCOUNT_NO', '') in self.REQUIRED_ACCOUNTS]
        
        logger.info(f"Filtered data size: {len(filtered_data)}")
        return filtered_data

    def _convert_numeric_columns(self, data: List[Dict[str, Any]], errors: List[str]) -> List[Dict[str, Any]]:
        """
        Converts numeric columns to Decimal.
        
        Args:
            data: List of transaction dictionaries
            errors: List to append conversion errors
            
        Returns:
            Data with numeric columns converted
        """
        logger.info(f"Converting numeric columns: {self.NUMERIC_COLUMNS}")
        
        conversion_errors = 0
        for row in data:
            for col in self.NUMERIC_COLUMNS:
                value = row.get(col, '')
                if value:
                    try:
                        row[col] = Decimal(str(value))
                    except (ValueError, Exception) as e:
                        logger.warning(f"Failed to convert {col} value '{value}' to Decimal: {str(e)}")
                        conversion_errors += 1
                        # Keep original value if conversion fails
        
        if conversion_errors > 0:
            logger.warning(f"Encountered {conversion_errors} conversion errors")
        
        return data

    def _group_by_date(self, data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Groups transactions by transaction date.
        
        Args:
            data: List of transaction dictionaries
            
        Returns:
            Dictionary mapping dates to lists of transactions
        """
        logger.info("Grouping transactions by date")
        
        # Sort by date first (required for itertools.groupby)
        sorted_data = sorted(data, key=lambda x: str(x.get('TRAN_DATE', '')))
        
        # Group by date
        date_groups = {}
        for date, group in itertools.groupby(sorted_data, key=lambda x: str(x.get('TRAN_DATE', ''))):
            date_groups[date] = list(group)
            logger.info(f"Date {date}: {len(date_groups[date])} transactions")
        
        logger.info(f"Grouped into {len(date_groups)} date groups")
        return date_groups

    def _create_account_excel(self, account_no: str, date: str, account_data: List[Dict[str, Any]]) -> bytes:
        """
        Creates an Excel file for a specific account/date combination.
        
        Args:
            account_no: Account number
            date: Transaction date
            account_data: List of transactions for this account/date
            
        Returns:
            Bytes content of the created Excel file
        """
        logger.info(f"Creating Excel file for account {account_no}, date {date}")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheet named by account number
        ws = wb.create_sheet(title=account_no)
        
        # Write headers
        ws.append(self.REQUIRED_CSV_COLUMNS)
        
        # Write data rows
        for row in account_data:
            row_values = [row.get(col, '') for col in self.REQUIRED_CSV_COLUMNS]
            ws.append(row_values)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Created Excel file with {len(account_data)} rows for account {account_no}, date {date}")
        return output.getvalue()

    def _create_summary_excel(self, date: str, date_data: List[Dict[str, Any]]) -> bytes:
        """
        Creates a summary Excel file with all accounts for a specific date.
        Each account gets its own sheet.
        
        Args:
            date: Transaction date
            date_data: All transactions for this date
            
        Returns:
            Bytes content of the created Excel file
        """
        logger.info(f"Creating summary Excel file for date {date}")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Group by account within this date
        sorted_by_account = sorted(date_data, key=lambda x: x.get('ACCOUNT_NO', ''))
        account_groups = itertools.groupby(sorted_by_account, key=lambda x: x.get('ACCOUNT_NO', ''))
        
        for account_no, account_data in account_groups:
            account_data = list(account_data)
            logger.info(f"Adding sheet for account {account_no} with {len(account_data)} transactions")
            
            # Create sheet named by account number
            ws = wb.create_sheet(title=account_no)
            
            # Write headers
            ws.append(self.REQUIRED_CSV_COLUMNS)
            
            # Write data rows
            for row in account_data:
                row_values = [row.get(col, '') for col in self.REQUIRED_CSV_COLUMNS]
                ws.append(row_values)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Created summary Excel file for date {date}")
        return output.getvalue()

    def _create_zip_file(self, excel_files: Dict[str, bytes], original_filename: str) -> FileModel:
        """
        Creates a ZIP file containing all the generated Excel files.
        
        Args:
            excel_files: Dictionary mapping filenames to file content bytes
            original_filename: Original filename (without extension)
            
        Returns:
            FileModel containing the ZIP file
        """
        logger.info(f"Creating ZIP file with {len(excel_files)} Excel files")
        
        zip_output = io.BytesIO()
        
        with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for filename, content in excel_files.items():
                zipf.writestr(filename, content)
                logger.debug(f"Added '{filename}' to ZIP archive")
        
        zip_output.seek(0)
        
        # Generate ZIP filename
        zip_filename = f"[pygrays]{original_filename}-BankStatement.zip"
        logger.info(f"Created ZIP file '{zip_filename}' with {len(excel_files)} files")
        
        return FileModel(name=zip_filename, content=zip_output.getvalue())

    async def process_uploaded_file(self, csv_file: FileModel) -> ResponseBase:
        """
        Processes an uploaded CSV file, filters by accounts, groups by date, and generates Excel files.
        
        Args:
            csv_file: FileModel containing the CSV file
            
        Returns:
            ResponseBase object with success status and FileModel data (ZIP file)
        """
        logger.info("=== Starting process_uploaded_file ===")
        logger.info(f"Received CSV file: {csv_file.name} ({len(csv_file.content)} bytes)")
        
        errors = []
        response = ResponseBase(is_success=True)
        
        try:
            # Step 1: Validate CSV file
            if not self._validate_csv_file(csv_file, errors):
                self._handle_errors(errors, response)
                return response
            
            # Step 2: Load CSV data
            data = self._load_csv_data(csv_file, errors)
            if data is None:
                self._handle_errors(errors, response)
                return response
            
            if len(data) == 0:
                error_msg = "No data rows found in CSV file"
                logger.error(error_msg)
                errors.append(error_msg)
                self._handle_errors(errors, response)
                return response
            
            # Step 3: Filter by accounts
            filtered_data = self._filter_by_accounts(data, errors)
            
            if len(filtered_data) == 0:
                error_msg = "No transactions found for required accounts"
                logger.error(error_msg)
                errors.append(error_msg)
                self._handle_errors(errors, response)
                return response
            
            # Step 4: Convert numeric columns
            filtered_data = self._convert_numeric_columns(filtered_data, errors)
            
            # Step 5: Group by date
            date_groups = self._group_by_date(filtered_data)
            
            if len(date_groups) == 0:
                error_msg = "No date groups found"
                logger.error(error_msg)
                errors.append(error_msg)
                self._handle_errors(errors, response)
                return response
            
            # Step 6: Create Excel files for each date/account combination
            excel_files = {}
            original_filename = csv_file.name.rsplit('.', 1)[0]  # Remove extension
            
            for date, date_data in date_groups.items():
                logger.info(f"Processing date {date} with {len(date_data)} transactions")
                
                # Group by account within this date
                sorted_by_account = sorted(date_data, key=lambda x: str(x.get('ACCOUNT_NO', '')))
                account_groups = itertools.groupby(sorted_by_account, key=lambda x: str(x.get('ACCOUNT_NO', '')))
                
                for account_no, account_data in account_groups:
                    account_data = list(account_data)
                    
                    # Create individual account file
                    account_excel = self._create_account_excel(account_no, date, account_data)
                    account_filename = f"{account_no}_{date}.xlsx"
                    excel_files[account_filename] = account_excel
                    logger.info(f"Created individual file: {account_filename}")
                
                # Create summary file for this date
                summary_excel = self._create_summary_excel(date, date_data)
                summary_filename = f"ALL Westpac Accounts Bank Statements {date}.xlsx"
                excel_files[summary_filename] = summary_excel
                logger.info(f"Created summary file: {summary_filename}")
            
            # Step 7: Create ZIP file
            zip_file = self._create_zip_file(excel_files, original_filename)
            
            # Set the data in the response object
            response.data = zip_file
            logger.info(f"Completed processing bank statement, returning ZIP file '{zip_file.name}'")
            return response
            
        except Exception as e:
            error_message = f"Error processing file: {str(e)}"
            logger.error(error_message, exc_info=True)
            errors.append(error_message)
            self._handle_errors(errors, response)
            return response

