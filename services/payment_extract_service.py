import io
import logging
import os
import zipfile
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.file_model import FileModel
from models.response_base import ResponseBase
from services.multi_logging import LoggingService
from utils.excel_utilities import ExcelUtilities

# Initialize logger
logger = LoggingService().get_logger(__name__)


class PaymentExtractService:
    """
    Service for processing Payment Extract Excel files.
    Splits data by unique BusinessEntity values and returns ZIP with individual files.
    """

    REQUIRED_SHEET_NAME = "Payments Extract"
    REQUIRED_COLUMN_NAME = "BusinessEntity"
    BLANK_VALUE_REPLACEMENT = "Blank"

    def __init__(self):
        pass

    def _handle_errors(self, errors: List[str], response: ResponseBase) -> bool:
        """
        Helper method to check for errors and update the response accordingly.
        
        Args:
            errors: List of error messages to check.
            response: Response object to update if errors exist.
            
        Returns:
            True if errors exist, False otherwise.
        """
        if len(errors) > 0:
            logger.warning(f"Handling {len(errors)} errors in response")
            for idx, err in enumerate(errors):
                logger.error(f"Error {idx+1}: {err}")
            response.errors.extend(errors)
            response.is_success = False
            return True
        logger.debug("No errors to handle in response")
        return False

    def _validate_excel_file(self, excel_file: FileModel, errors: List[str]) -> Optional[openpyxl.Workbook]:
        """
        Validates the Excel file structure and loads the workbook.
        Supports both XLS and XLSX formats.
        
        Args:
            excel_file: FileModel containing the Excel file
            errors: List to append validation errors
            
        Returns:
            Loaded workbook if valid, None otherwise
        """
        logger.info(f"Validating Excel file: {excel_file.name}")
        
        # Use shared utility to load Excel file (supports both XLS and XLSX)
        wb = ExcelUtilities.load_excel_workbook(excel_file, errors=errors, read_only=False, data_only=True)  # type: ignore[attr-defined]
        
        if wb is None:
            return None
        
        # Validate required sheet exists
        if self.REQUIRED_SHEET_NAME not in wb.sheetnames:
            error_msg = f"Sheet '{self.REQUIRED_SHEET_NAME}' not found in the Excel file"
            logger.error(error_msg)
            errors.append(error_msg)
            return None
        
        logger.info(f"Found required sheet '{self.REQUIRED_SHEET_NAME}'")
        return wb

    def _validate_sheet_structure(self, sheet: Worksheet, errors: List[str]) -> Optional[int]:
        """
        Validates that the sheet has the required BusinessEntity column.
        
        Args:
            sheet: The worksheet to validate
            errors: List to append validation errors
            
        Returns:
            Column index of BusinessEntity if found, None otherwise
        """
        logger.info(f"Validating sheet structure for '{sheet.title}'")
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            header_value = cell.value
            if header_value is not None:
                headers.append(str(header_value).strip())
            else:
                headers.append("")
        
        logger.debug(f"Found {len(headers)} columns in header row: {headers}")
        
        # Validate BusinessEntity column exists
        if self.REQUIRED_COLUMN_NAME not in headers:
            error_msg = f"Column '{self.REQUIRED_COLUMN_NAME}' not found in the '{self.REQUIRED_SHEET_NAME}' sheet"
            logger.error(error_msg)
            errors.append(error_msg)
            return None
        
        business_entity_col_idx = headers.index(self.REQUIRED_COLUMN_NAME) + 1  # openpyxl uses 1-based indexing
        logger.info(f"Found '{self.REQUIRED_COLUMN_NAME}' column at index {business_entity_col_idx}")
        
        return business_entity_col_idx

    def _read_sheet_data(self, sheet: Worksheet, 
                        business_entity_col_idx: int) -> List[Dict[str, Any]]:
        """
        Reads all data rows from the sheet into a list of dictionaries.
        
        Args:
            sheet: The worksheet to read from
            business_entity_col_idx: Column index (1-based) of BusinessEntity
            
        Returns:
            List of dictionaries, each representing a row
        """
        logger.info(f"Reading data from sheet '{sheet.title}'")
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            header_value = cell.value
            if header_value is not None:
                headers.append(str(header_value).strip())
            else:
                headers.append("")
        
        data_rows = []
        row_count = 0
        
        # Read all data rows (skip header row)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            row_count += 1
            row_dict = {}
            
            for col_idx, cell in enumerate(row, start=1):
                if col_idx <= len(headers):
                    header_name = headers[col_idx - 1]
                    cell_value = cell.value
                    
                    # Convert None to empty string for consistency
                    if cell_value is None:
                        cell_value = ""
                    
                    row_dict[header_name] = cell_value
            
            data_rows.append(row_dict)
        
        logger.info(f"Read {row_count} data rows from sheet '{sheet.title}'")
        return data_rows

    def _get_unique_business_entities(self, data_rows: List[Dict[str, Any]], 
                                      business_entity_col_name: str) -> List[str]:
        """
        Extracts unique BusinessEntity values from the data.
        Empty strings and None values are treated as "Blank".
        
        Args:
            data_rows: List of row dictionaries
            business_entity_col_name: Name of the BusinessEntity column
            
        Returns:
            List of unique BusinessEntity values (with "Blank" for empty values)
        """
        logger.info(f"Extracting unique BusinessEntity values from {len(data_rows)} rows")
        
        unique_values = set()
        
        for row in data_rows:
            business_entity_value = row.get(business_entity_col_name, "")
            
            # Convert None to empty string
            if business_entity_value is None:
                business_entity_value = ""
            
            # Convert to string and strip whitespace
            business_entity_value = str(business_entity_value).strip()
            
            # Replace empty string with "Blank"
            if business_entity_value == "":
                business_entity_value = self.BLANK_VALUE_REPLACEMENT
            
            unique_values.add(business_entity_value)
        
        unique_list = sorted(list(unique_values))
        logger.info(f"Found {len(unique_list)} unique BusinessEntity values: {unique_list}")
        
        return unique_list

    def _create_excel_file_for_entity(self, data_rows: List[Dict[str, Any]], 
                                      business_entity_value: str,
                                      business_entity_col_name: str,
                                      headers: List[str]) -> bytes:
        """
        Creates an Excel file containing only rows for a specific BusinessEntity value.
        
        Args:
            data_rows: All data rows
            business_entity_value: The BusinessEntity value to filter by
            business_entity_col_name: Name of the BusinessEntity column
            headers: List of column headers
            
        Returns:
            Bytes content of the created Excel file
        """
        logger.info(f"Creating Excel file for BusinessEntity: '{business_entity_value}'")
        
        # Filter rows for this BusinessEntity
        filtered_rows = []
        for row in data_rows:
            row_entity_value = row.get(business_entity_col_name, "")
            
            # Convert None to empty string
            if row_entity_value is None:
                row_entity_value = ""
            
            # Convert to string and strip
            row_entity_value = str(row_entity_value).strip()
            
            # Replace empty string with "Blank" for comparison
            if row_entity_value == "":
                row_entity_value = self.BLANK_VALUE_REPLACEMENT
            
            if row_entity_value == business_entity_value:
                filtered_rows.append(row)
        
        logger.info(f"Filtered {len(filtered_rows)} rows for BusinessEntity '{business_entity_value}'")
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = self.REQUIRED_SHEET_NAME
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data rows
        for row_idx, row_dict in enumerate(filtered_rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell_value = row_dict.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Created Excel file with {len(filtered_rows)} rows for BusinessEntity '{business_entity_value}'")
        return output.getvalue()

    def _create_zip_file(self, excel_files: Dict[str, bytes], original_filename: str) -> FileModel:
        """
        Creates a ZIP file containing all the generated Excel files.
        
        Args:
            excel_files: Dictionary mapping filenames to file content bytes
            original_filename: Original filename (without extension)
            
        Returns:
            FileModel containing the ZIP file
        """
        logger.info(f"Creating ZIP file with {len(excel_files)} Excel files")
        
        zip_output = io.BytesIO()
        
        with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for filename, content in excel_files.items():
                zipf.writestr(filename, content)
                logger.debug(f"Added '{filename}' to ZIP archive")
        
        zip_output.seek(0)
        
        # Generate ZIP filename
        zip_filename = f"[pygrays]{original_filename}-PaymentExtract.zip"
        logger.info(f"Created ZIP file '{zip_filename}' with {len(excel_files)} files")
        
        return FileModel(name=zip_filename, content=zip_output.getvalue())

    async def process_uploaded_file(self, excel_file: FileModel) -> ResponseBase:
        """
        Processes an uploaded Excel file, splits by BusinessEntity, and returns ZIP.
        
        Args:
            excel_file: FileModel containing the Excel file
            
        Returns:
            ResponseBase object with success status and FileModel data (ZIP file)
        """
        logger.info("=== Starting process_uploaded_file ===")
        logger.info(f"Received Excel file: {excel_file.name} ({len(excel_file.content)} bytes)")
        
        errors = []
        response = ResponseBase(is_success=True)
        
        try:
            # Step 1: Validate and load Excel file
            wb = self._validate_excel_file(excel_file, errors)
            if wb is None:
                self._handle_errors(errors, response)
                return response
            
            # Step 2: Get the required sheet
            sheet = wb[self.REQUIRED_SHEET_NAME]
            logger.info(f"Accessing sheet '{self.REQUIRED_SHEET_NAME}'")
            
            # Step 3: Validate sheet structure
            business_entity_col_idx = self._validate_sheet_structure(sheet, errors)
            if business_entity_col_idx is None:
                self._handle_errors(errors, response)
                return response
            
            # Step 4: Read all data from the sheet
            data_rows = self._read_sheet_data(sheet, business_entity_col_idx)
            
            if len(data_rows) == 0:
                error_msg = f"No data rows found in sheet '{self.REQUIRED_SHEET_NAME}'"
                logger.error(error_msg)
                errors.append(error_msg)
                self._handle_errors(errors, response)
                return response
            
            # Step 5: Get headers for creating new files
            headers = []
            for cell in sheet[1]:
                header_value = cell.value
                if header_value is not None:
                    headers.append(str(header_value).strip())
                else:
                    headers.append("")
            
            # Step 6: Get unique BusinessEntity values
            unique_entities = self._get_unique_business_entities(data_rows, self.REQUIRED_COLUMN_NAME)
            
            if len(unique_entities) == 0:
                error_msg = "No unique BusinessEntity values found"
                logger.error(error_msg)
                errors.append(error_msg)
                self._handle_errors(errors, response)
                return response
            
            # Step 7: Create Excel file for each unique BusinessEntity
            excel_files = {}
            original_filename = os.path.splitext(excel_file.name)[0]  # Remove extension
            
            for entity_value in unique_entities:
                # Create XLSX file first
                xlsx_content = self._create_excel_file_for_entity(
                    data_rows, 
                    entity_value, 
                    self.REQUIRED_COLUMN_NAME,
                    headers
                )
                
                # Convert XLSX to XLS
                xls_content = ExcelUtilities.convert_xlsx_to_xls(xlsx_content, f"{original_filename}-BusinessEntity-{entity_value}.xlsx", errors)  # type: ignore[attr-defined]
                if xls_content is None:
                    error_msg = f"Failed to convert XLSX to XLS for BusinessEntity '{entity_value}'"
                    logger.error(error_msg)
                    errors.append(error_msg)
                    self._handle_errors(errors, response)
                    return response
                
                # Generate filename with .xls extension
                excel_filename = f"{original_filename}-BusinessEntity-{entity_value}.xls"
                excel_files[excel_filename] = xls_content
                logger.info(f"Created and converted Excel file to XLS: {excel_filename}")
            
            # Step 8: Create ZIP file
            zip_file = self._create_zip_file(excel_files, original_filename)
            
            # Set the data in the response object
            response.data = zip_file
            logger.info(f"Completed processing payment extract, returning ZIP file '{zip_file.name}'")
            return response
            
        except Exception as e:
            error_message = f"Error processing file: {str(e)}"
            logger.error(error_message, exc_info=True)
            errors.append(error_message)
            self._handle_errors(errors, response)
            return response


