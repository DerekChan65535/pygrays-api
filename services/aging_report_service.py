import csv
import io
import logging
import re
import time
from datetime import datetime, timedelta
from typing import List, Optional, Tuple, Dict, Any, Union
import zipfile

import openpyxl
from openpyxl.styles import PatternFill

from models.file_model import FileModel
from models.response_base import ResponseBase
from utils.schema_config import (
    aging_report_daily_data_import_schema, 
    BaseSchema, 
    ExportSchema, 
    ImportSchema, 
    aging_report_data_schema,
    aging_report_fully_paid_schema,
    aging_report_not_fully_paid_schema,
    ImportField, 
    ExportField
)

# Initialize logger with detailed configuration
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class AgingReportService:
    # Use the schema from configuration
    daily_data_import_schema: ImportSchema = aging_report_daily_data_import_schema


    @staticmethod
    def parse_date_with_formats(date_string: str, formats: List[str]) -> Optional[datetime]:
        """
        Try parsing a date string using multiple formats
        
        Args:
            date_string: The date string to parse
            formats: List of format strings to try
            
        Returns:
            Parsed datetime object or None if parsing fails
        """
        if not date_string:
            logger.debug(f"Empty date string provided, returning None")
            return None

        # Ensure formats is a list
        if isinstance(formats, str):
            logger.debug(f"Converting single format '{formats}' to list")
            formats = [formats]

        # Try each format
        for fmt in formats:
            try:
                parsed_date = datetime.strptime(date_string, fmt)
                logger.debug(f"Successfully parsed date '{date_string}' with format '{fmt}'")
                return parsed_date
            except ValueError:
                logger.debug(f"Failed to parse date '{date_string}' with format '{fmt}'")
                continue

        # If all formats fail, return None
        logger.warning(f"Failed to parse date '{date_string}' with any provided formats: {formats}")
        return None

    def __init__(self):
        pass

    def _handle_errors(self, errors: List[str], response: ResponseBase) -> bool:
        """
        Helper method to check for errors and update the response accordingly.
        
        Args:
            errors: List of error messages to check.
            response: Response object to update if errors exist.
            
        Returns:
            True if errors exist, False otherwise.
        """
        if len(errors) > 0:
            logger.warning(f"Handling {len(errors)} errors in response")
            for idx, err in enumerate(errors):
                logger.error(f"Error {idx+1}: {err}")
            response.errors.extend(errors)
            response.is_success = False
            return True
        logger.debug("No errors to handle in response")
        return False

    def _load_and_process_mapping_file(
        self, mapping_file: FileModel, errors: List[str]
    ) -> Optional[Tuple[List[Dict[str, str]], List[Dict[str, str]], List[Dict[str, Union[str, int, None]]]]]:
        """
        Loads and processes the mapping file to create lookup dictionaries.
        Returns a tuple of dictionaries or None if a critical error occurs.
        """
        logger.info(f"Processing mapping file: {mapping_file.name}")

        try:
            # Decode using UTF-8-SIG encoding to remove BOM if present
            tables_data_str = mapping_file.content.decode('utf-8-sig')
            logger.debug(f"Successfully decoded mapping file content ({len(tables_data_str)} characters)")
            tables_data_reader = csv.reader(io.StringIO(tables_data_str))
            logger.debug("Created CSV reader for mapping file to access columns by index")
        except Exception as decode_error:
            error_msg = f"Error decoding mapping file {mapping_file.name}: {str(decode_error)}"
            logger.error(error_msg, exc_info=True)
            errors.append(error_msg)
            return None

        division_to_subdivision: List[Dict[str, str]] = []
        divisionno_to_division: List[Dict[str, str]] = []
        division_state_days: List[Dict[str, Union[str, int, None]]] = []
        
        mapping_errors_count = 0  # For potential future use to count row-specific parsing errors

        logger.info("Building lookup dictionaries from mapping file using column indices")
        tables_data = list(tables_data_reader)

        if not tables_data:
            error_msg = f"Mapping file {mapping_file.name} is empty or contains no header row."
            logger.error(error_msg)
            errors.append(error_msg)
            return None
            
        mapping_file_headers = tables_data[0]
        data_rows = tables_data[1:] # Skip header row for data processing

        # The col 0 and 1 of the mapping file is for Division and Sub Division
        try:
            div_subdiv_header_indices: Tuple[int, int] = (0, 1)
            division_subdivision_headers: List[str] = [mapping_file_headers[i] for i in div_subdiv_header_indices]
            division_to_subdivision_raw: List[List[str]] = [[row[i] for i in div_subdiv_header_indices if i < len(row)] for row in data_rows]
            division_to_subdivision: List[Dict[str, str]] = [
                dict(zip(division_subdivision_headers, row_values))
                for row_values in division_to_subdivision_raw
                if len(row_values) == len(division_subdivision_headers) and all(row_values)
            ]
        except IndexError:
            errors.append(f"Mapping file {mapping_file.name} has insufficient columns for Division/Sub Division mapping.")
            logger.warning(f"Skipping Division/Sub Division mapping due to insufficient columns in {mapping_file.name}.")


        # The col 3 and 4 of the mapping file is for DivisionNo and Division
        try:
            divno_div_header_indices: Tuple[int, int] = (3, 4)
            divisionno_division_headers: List[str] = [mapping_file_headers[i] for i in divno_div_header_indices]
            divisionno_to_division_raw: List[List[str]] = [[row[i] for i in divno_div_header_indices if i < len(row)] for row in data_rows]
            divisionno_to_division: List[Dict[str, str]] = [
                dict(zip(divisionno_division_headers, row_values))
                for row_values in divisionno_to_division_raw
                if len(row_values) == len(divisionno_division_headers) and all(row_values)
            ]
        except IndexError:
            errors.append(f"Mapping file {mapping_file.name} has insufficient columns for DivisionNo/Division mapping.")
            logger.warning(f"Skipping DivisionNo/Division mapping due to insufficient columns in {mapping_file.name}.")

        # The col 6, 7, 9 of the mapping file is for Division Name, State, Days
        try:
            div_state_days_indices: Tuple[int, int, int] = (6, 7, 9)
            division_state_days_headers: List[str] = [mapping_file_headers[i] for i in div_state_days_indices]
            
            for row in data_rows:
                if all(i < len(row) for i in div_state_days_indices): # Ensure all indices are within row bounds
                    raw_row_list = [row[div_state_days_indices[0]], row[div_state_days_indices[1]], row[div_state_days_indices[2]]]
                    entry: Dict[str, Union[str, int, None]] = dict(zip(division_state_days_headers, raw_row_list))
                    try:
                        days_value_str = str(entry.get("Days", "")).strip()
                        if days_value_str:
                            entry["Days"] = int(days_value_str)
                        else:
                            entry["Days"] = None 
                    except ValueError:
                        logger.warning(f"Could not convert 'Days' value '{entry.get('Days')}' to int for entry: {entry} in {mapping_file.name}")
                        entry["Days"] = None 
                    
                    if entry.get(mapping_file_headers[div_state_days_indices[1]]) and entry.get(mapping_file_headers[div_state_days_indices[0]]): # Check using actual header names for State and Division Name
                         division_state_days.append(entry)
                else:
                    logger.warning(f"Skipping mapping row in {mapping_file.name} due to insufficient columns for Division/State/Days: {row}")
        except IndexError:
            errors.append(f"Mapping file {mapping_file.name} has insufficient columns for Division/State/Days mapping.")
            logger.warning(f"Skipping Division/State/Days mapping due to insufficient columns in {mapping_file.name}.")


        logger.info(f"Completed processing {len(tables_data)} rows from mapping file: {mapping_file.name}")
        logger.info(f"Created lookup dictionaries: division_to_subdivision ({len(division_to_subdivision)} entries), "
                   f"divisionno_to_division ({len(divisionno_to_division)} entries), "
                   f"division_state_days ({len(division_state_days)} entries)")

        if mapping_errors_count > 0: # This count is not currently incremented but is here for structure
            logger.warning(f"Encountered {mapping_errors_count} issues while processing mapping file rows from {mapping_file.name}.")

        return division_to_subdivision, divisionno_to_division, division_state_days

    def _validate_and_extract_file_info(self, data_files: List['FileModel'], errors: List[str]) -> List[Tuple[str, 'FileModel']]:
        """
        Validates data files and extracts state information from filenames.
        
        Args:
            data_files: List of data files to validate
            errors: List to append any validation errors
            
        Returns:
            List of tuples containing (state, file) for valid files
        """
        if not data_files or len(data_files) == 0:
            error_msg = "No data files provided"
            errors.append(error_msg)
            logger.error(error_msg)
            return []

        valid_files = []
        logger.info(f"Validating {len(data_files)} data files")

        for file_idx, data_file in enumerate(data_files, 1):
            logger.info(f"Validating file {file_idx}/{len(data_files)}: {data_file.name}")

            # Extract state from filename using regex pattern
            state_match = re.search(r'(?:Sales[ _]?Aged[ _]?Balance[ _]?(?:\s*-\s*)?|SalesAgedBalance)(\w+)\.csv', data_file.name, re.IGNORECASE)
            if not state_match:
                error_msg = (f"Unable to extract state from filename: {data_file.name}. Expected formats: "
                           f"'Sales Aged Balance [state].csv', 'SalesAgedBalance[state].csv', 'Sales_Aged_Balance_[state].csv', or 'Sales Aged Balance - [state].csv'")
                errors.append(error_msg)
                logger.error(error_msg)
                continue

            state = state_match.group(1).upper()
            logger.info(f"Extracted state '{state}' from filename {data_file.name}")
            valid_files.append((state, data_file))

        return valid_files

    def _load_and_filter_data_files(self, file_info_list: List[Tuple[str, 'FileModel']], errors: List[str]) -> List[Dict[str, Any]]:
        """
        Loads and filters data from all valid data files.
        
        Args:
            file_info_list: List of (state, file) tuples
            errors: List to append any processing errors
            
        Returns:
            Combined filtered data from all files
        """
        all_filtered_data = []
        
        for state, data_file in file_info_list:
            file_start_time = time.time()
            logger.info(f"Processing data file: {data_file.name} (State: {state})")

            try:
                # Load data using schema
                daily_data = self.daily_data_import_schema.import_data(data_file.content, errors)
                logger.info(f'Loaded {len(daily_data)} rows from {data_file.name}')

                # Filter data based on business rules
                filtered_data = self._apply_data_filters(daily_data, state)
                logger.info(f"Filtered to {len(filtered_data)} rows for {data_file.name}")
                
                all_filtered_data.extend(filtered_data)
                
            except Exception as e:
                error_msg = f"Error processing file {data_file.name}: {str(e)}"
                logger.error(error_msg, exc_info=True)
                errors.append(error_msg)
                continue

        return all_filtered_data

    def _apply_data_filters(self, daily_data: List[Dict[str, Any]], state: str) -> List[Dict[str, Any]]:
        """
        Applies business rule filters to the daily data.
        
        Args:
            daily_data: Raw data from the CSV file
            state: State code to add to each row
            
        Returns:
            Filtered data with state information added
        """
        filtered_data = []
        excluded_count = {
            "cheque_date": 0,
            "zero_gross": 0,
            "cancellation": 0,
            "totals_rows": 0
        }

        logger.info(f"Applying filters to {len(daily_data)} rows")

        for row_idx, row_dict in enumerate(daily_data):
            # Extract key fields for filtering
            cheque_date = row_dict.get('Cheque_Date')
            gross_total = row_dict.get('Gross_Tot')
            description = row_dict.get('Description')

            # Apply exclusion rules
            if cheque_date is not None:
                excluded_count["cheque_date"] += 1
                logger.debug(f"Excluding row {row_idx}: Non-null Cheque_Date={cheque_date}")
                continue
                
            if gross_total == 0:
                excluded_count['zero_gross'] += 1
                logger.debug(f'Excluding row {row_idx}: Zero Gross_Tot')
                continue
                
            if description and "Buyer Cancellation Fees" in str(description):
                excluded_count["cancellation"] += 1
                logger.debug(f"Excluding row {row_idx}: Buyer Cancellation Fees in description")
                continue
                
            if description and any(total_text in str(description) for total_text in ['Total Invoices', 'Total Payments', 'Total Bankings']):
                excluded_count["totals_rows"] += 1
                logger.debug(f"Excluding row {row_idx}: Found totals text in description: '{description}'")
                continue

            # Add state and include row
            row_dict['State'] = state
            filtered_data.append(row_dict)

        logger.info(f"Filtering complete. Kept {len(filtered_data)} rows, excluded {len(daily_data) - len(filtered_data)} rows")
        logger.info(f"Exclusion breakdown: {excluded_count}")
        
        return filtered_data

    def _transform_data_rows(self, filtered_data: List[Dict[str, Any]], mapping_data: Tuple, reporting_date: datetime, errors: List[str]) -> List[Dict[str, Any]]:
        """
        Transforms filtered data by computing new columns based on business logic.
        
        Args:
            filtered_data: Filtered data from all files
            mapping_data: Tuple containing mapping dictionaries
            reporting_date: Date for calculations
            errors: List to append any transformation errors
            
        Returns:
            Transformed data with computed columns
        """
        division_to_subdivision, divisionno_to_division, division_state_days = mapping_data
        transformed_rows = []
        mapping_errors = []

        logger.info(f"Transforming {len(filtered_data)} rows")

        for row_dict in filtered_data:
            new_row = row_dict.copy()
            
            if not row_dict.get('Classification'):
                logger.debug(f"Skipping row without Classification: {row_dict.get('Sale_No', 'Unknown')}")
                transformed_rows.append(new_row)
                continue

            try:
                self._compute_derived_columns(new_row, row_dict, division_to_subdivision, 
                                            divisionno_to_division, division_state_days, 
                                            reporting_date, mapping_errors)
                transformed_rows.append(new_row)
                
            except Exception as e:
                error_msg = f"Error transforming row {row_dict.get('Sale_No', 'Unknown')}: {str(e)}"
                logger.error(error_msg, exc_info=True)
                mapping_errors.append(error_msg)

        # Handle mapping errors
        if mapping_errors:
            logger.error(f"Found {len(mapping_errors)} transformation errors")
            max_errors_to_show = 10
            if len(mapping_errors) > max_errors_to_show:
                truncated_msg = f"Found {len(mapping_errors)} transformation errors. First {max_errors_to_show} errors shown:"
                errors.append(truncated_msg)
                errors.extend(mapping_errors[:max_errors_to_show])
                errors.append(f"...and {len(mapping_errors) - max_errors_to_show} more errors.")
            else:
                errors.extend(mapping_errors)

        logger.info(f"Completed transformation of {len(transformed_rows)} rows")
        return transformed_rows

    def _compute_derived_columns(self, new_row: Dict[str, Any], row_dict: Dict[str, Any], 
                                division_to_subdivision: List[Dict[str, str]], 
                                divisionno_to_division: List[Dict[str, str]], 
                                division_state_days: List[Dict[str, Union[str, int, None]]], 
                                reporting_date: datetime, mapping_errors: List[str]) -> None:
        """
        Computes all derived columns (43-55) for a single row.
        
        Args:
            new_row: Row dictionary to update with computed values
            row_dict: Original row data
            division_to_subdivision: Division to subdivision mapping
            divisionno_to_division: Division number to division mapping  
            division_state_days: State/division to payment days mapping
            reporting_date: Date for calculations
            mapping_errors: List to append mapping errors
        """
        # Column 46 (AT): Lookup Division from DivisionNo
        division_no = row_dict.get('Division')
        division_entry = next((item for item in divisionno_to_division if item.get("DivisionNo") == division_no), None)
        if not division_entry:
            error_msg = f"Missing Division mapping for DivisionNo '{division_no}' in Sale_No {row_dict.get('Sale_No', 'Unknown')}"
            logger.error(error_msg)
            mapping_errors.append(error_msg)
        
        new_row['Division Name'] = division_entry.get("Division", "") if division_entry else ""

        # Column 43 (AQ): Concatenate State and Division Name
        state_val = row_dict.get('State') or ""
        new_row['State-Division Name'] = f"{state_val}-{new_row['Division Name']}" if state_val and new_row['Division Name'] else ""

        # Column 44 (AR): Lookup Payment Days
        state_division = new_row['State-Division Name']
        state_days_entry = next((item for item in division_state_days 
                               if f"{item.get('State')}-{item.get('Division Name')}" == state_division), None)
        if not state_days_entry and state_division:
            error_msg = f"Missing Payment Days mapping for State-Division '{state_division}' in Sale_No {row_dict.get('Sale_No', 'Unknown')}"
            logger.error(error_msg)
            mapping_errors.append(error_msg)
        
        new_row['Payment Days'] = state_days_entry.get("Days", "") if state_days_entry else ""

        # Column 45 (AS): Calculate Due Date
        sale_date = row_dict.get('Sale_Date')
        payment_days = new_row['Payment Days']
        if isinstance(sale_date, datetime) and isinstance(payment_days, int):
            new_row['Due Date'] = sale_date + timedelta(days=payment_days)
        else:
            new_row['Due Date'] = ""

        # Column 47 (AU): Lookup Sub Division
        division = new_row['Division Name']
        subdivision_entry = next((item for item in division_to_subdivision 
                                if item.get("Division") == division), None)
        if not subdivision_entry and division:
            error_msg = f"Missing Sub Division mapping for Division '{division}' in Sale_No {row_dict.get('Sale_No', 'Unknown')}"
            logger.error(error_msg)
            mapping_errors.append(error_msg)
        
        new_row['Sub Division Name'] = subdivision_entry.get("Sub Division", "") if subdivision_entry else ""

        # Column 48 (AV): Compute Gross Amount
        delot_ind = str(row_dict.get('Delot_Ind', "")).upper() == "TRUE"
        gross_tot = row_dict.get('Gross_Tot')
        sale_no = row_dict.get('Sale_No')
        
        if delot_ind and isinstance(gross_tot, (int, float)) and isinstance(sale_no, (int, float)):
            current_gross_amount_num = float(gross_tot - sale_no)
        elif isinstance(gross_tot, (int, float)):
            current_gross_amount_num = float(gross_tot)
        else:
            current_gross_amount_num = 0.0
        new_row['Gross Amount'] = current_gross_amount_num

        # Column 50 (AX): Get To be Collected
        day_num = reporting_date.day
        day_key = f"Day{day_num}"
        source_tbc_val = row_dict.get(day_key, None)
        numeric_to_be_collected = float(source_tbc_val) if isinstance(source_tbc_val, (int, float)) else 0.0
        new_row['To be Collected'] = numeric_to_be_collected

        # Column 49 (AW): Calculate Collected
        new_row['Collected'] = current_gross_amount_num - numeric_to_be_collected

        # Column 51 (AY): Compute Payable to Vendor
        payable_to_vendor_val_num = 0.0
        if delot_ind and numeric_to_be_collected == 0.0:
            payable_to_vendor_val_num = current_gross_amount_num
        new_row['Payable to Vendor'] = payable_to_vendor_val_num

        # Column 52 (AZ): Format Month
        if row_dict.get('Description') and isinstance(sale_date, datetime):
            new_row['Month'] = sale_date.strftime("%b-%y")
        else:
            new_row['Month'] = ""

        # Column 53 (BA): Extract Year
        if row_dict.get('Description') and isinstance(sale_date, datetime):
            new_row['Year'] = sale_date.year
        else:
            new_row['Year'] = ""

        # Column 54 (BB): Cheque Date Y/N
        new_row['Cheque Date Y/N'] = "YES" if row_dict.get('Cheque_Date') else "NO"

        # Column 55 (BC): Compute days late
        try:
            if new_row['Payable to Vendor'] == row_dict.get('Gross_Tot'):
                cheque_date_val = row_dict.get('Cheque_Date')
                if isinstance(cheque_date_val, datetime):
                    days_diff = (reporting_date.date() - cheque_date_val.date()).days
                    new_row['Days Late for Vendors Pmt'] = days_diff if days_diff > 0 else ''
                else:
                    new_row['Days Late for Vendors Pmt'] = ''
            else:
                new_row['Days Late for Vendors Pmt'] = ''
        except:
            new_row['Days Late for Vendors Pmt'] = ''

    def _create_excel_reports(self, processed_data: List[Dict[str, Any]], mapping_file: 'FileModel', 
                             date_str: str, reporting_date: datetime, errors: List[str]) -> 'FileModel':
        """
        Creates Excel reports and packages them into a ZIP file.
        
        Args:
            processed_data: Transformed data ready for export
            mapping_file: Original mapping file for Tables sheet
            date_str: Formatted date string for filenames
            reporting_date: Date for conditional formatting
            errors: List to append any export errors
            
        Returns:
            FileModel containing the ZIP file with all reports
        """
        logger.info(f"Creating Excel reports for {len(processed_data)} rows")

        # Create main workbook with Tables sheet
        template_wb = openpyxl.Workbook()
        if template_wb.active:
            template_wb.remove(template_wb.active)

        # Add Tables sheet with mapping data
        self._create_tables_sheet(template_wb, mapping_file)

        # Export main data
        errors_export = []
        success = aging_report_data_schema.export_data(processed_data, template_wb, '---DATA---', errors_export)
        if not success:
            logger.error(f"Failed to export data to '---DATA---' sheet: {errors_export}")
            errors.extend(errors_export)
            raise Exception("Failed to export main data")

        # Create ZIP file with all reports
        zip_output = io.BytesIO()
        with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add main Excel file
            output = io.BytesIO()
            template_wb.save(output)
            output.seek(0)
            zipf.writestr(f"Sales_Aged_Balance_Report_{date_str}.xlsx", output.getvalue())

            # Create filtered reports
            self._create_filtered_reports(zipf, processed_data, date_str, reporting_date, errors)

        zip_output.seek(0)
        zip_file_name = f"[pygrays]Sales_Aged_Balance_Reports_{date_str}.zip"
        
        logger.info(f"Created ZIP file '{zip_file_name}' with all reports")
        return FileModel(name=zip_file_name, content=zip_output.getvalue())

    def _create_tables_sheet(self, workbook: openpyxl.Workbook, mapping_file: 'FileModel') -> None:
        """Creates the Tables sheet with mapping data."""
        tables_sheet = workbook.create_sheet(title='Tables')
        logger.debug("Added 'Tables' sheet to output workbook")

        tables_data_str = mapping_file.content.decode('utf-8')
        tables_data_csv = csv.reader(io.StringIO(tables_data_str))
        tables_data_rows = list(tables_data_csv)

        for row_idx, row in enumerate(tables_data_rows, 1):
            for col_idx, value in enumerate(row, 1):
                tables_sheet.cell(row=row_idx, column=col_idx).value = value

    def _create_filtered_reports(self, zipf: zipfile.ZipFile, processed_data: List[Dict[str, Any]], 
                                date_str: str, reporting_date: datetime, errors: List[str]) -> None:
        """Creates filtered Excel reports for different divisions."""
        filter_criteria = {
            'DivAUTO': ["BANKING, INSOLVENCY & FINANCE", "CONSUMER", "INDUSTRIAL", "WINE"],
            'DivINDUSTRIAL': ["AUTO W", "CONSUMER", "BOATS", "CARAVANS", "WINE"],
            'DivCONSUMER': ["AUTO", "BANKING, INSOLVENCY & FINANCE", "BOATS", "CARAVANS", "INDUSTRIAL", "WINE"]
        }

        yesterday = reporting_date.date() - timedelta(days=1)
        context = {'yesterday': yesterday}

        for report_type, criteria in filter_criteria.items():
            logger.info(f"Creating Excel file for {report_type}")
            
            # Filter data
            filtered_data = [row for row in processed_data if row.get('Sub Division Name') in criteria]
            logger.info(f"Filtered {len(filtered_data)} rows for {report_type}")

            # Create workbook
            report_wb = openpyxl.Workbook()
            if report_wb.active:
                report_wb.remove(report_wb.active)

            # Create sheets
            errors_export = []
            
            # FULLY PAID sheet
            fully_paid_data = [row for row in filtered_data if row.get('To be Collected') == 0.0]
            success = aging_report_fully_paid_schema.export_data(fully_paid_data, report_wb, 'FULLY PAID', errors_export, context)
            if not success:
                logger.error(f"Failed to export FULLY PAID data for {report_type}: {errors_export}")
                errors.extend(errors_export)
                continue

            # NOT FULLY PAID sheet
            not_fully_paid_data = [row for row in filtered_data if row.get('To be Collected') != 0.0]
            success = aging_report_not_fully_paid_schema.export_data(not_fully_paid_data, report_wb, 'NOT FULLY PAID', errors_export)
            if not success:
                logger.error(f"Failed to export NOT FULLY PAID data for {report_type}: {errors_export}")
                errors.extend(errors_export)
                continue

            # Save to ZIP
            report_output = io.BytesIO()
            report_wb.save(report_output)
            report_output.seek(0)

            # Determine filename
            file_names = {
                'DivAUTO': f"All Sales Aged Balance Report {date_str} - Auto.xlsx",
                'DivINDUSTRIAL': f"All Sales Aged Balance Report {date_str} - Industrial.xlsx",
                'DivCONSUMER': f"All Sales Aged Balance Report {date_str} - Consumer.xlsx"
            }
            
            zipf.writestr(file_names[report_type], report_output.getvalue())

    async def process_uploaded_file(self, mapping_file: 'FileModel',
                                   data_files: List['FileModel'],
                                   report_date: datetime) -> 'ResponseBase':
        """
        Processes multiple daily Sales Aged Balance reports, computes values for columns 43 to 55,
        and returns a FileModel with the combined processed data.

        Args:
            mapping_file: CSV file containing mapping tables
            data_files: List of CSV files containing daily sales data with state info in filenames
            report_date: Specific date to use for report calculations

        Returns:
            ResponseBase object with success status and FileModel data
        """
        method_start_time = time.time()
        logger.info("=== Starting process_uploaded_file ===")
        logger.info(f"Received mapping file: {mapping_file.name} ({len(mapping_file.content)} bytes)")
        logger.info(f"Received {len(data_files)} data files")

        errors = []
        response = ResponseBase(is_success=True)

        try:
            # Use the provided report_date
            reporting_date = report_date
            date_str = reporting_date.strftime("%Y%m%d")
            logger.debug(f"Processing date: {reporting_date}, formatted as {date_str}")

            # Step 1: Validate files and extract state information
            valid_file_info = self._validate_and_extract_file_info(data_files, errors)
            if self._handle_errors(errors, response):
                return response

            # Step 2: Load and filter data from all files
            all_filtered_data = self._load_and_filter_data_files(valid_file_info, errors)
            if self._handle_errors(errors, response):
                return response

            if len(all_filtered_data) == 0:
                error_msg = f"No data was extracted from any of the {len(data_files)} data files."
                logger.error(error_msg)
                errors.append(error_msg)
                self._handle_errors(errors, response)
                return response

            # Step 3: Load mapping file
            mapping_data = self._load_and_process_mapping_file(mapping_file, errors)
            if mapping_data is None:
                self._handle_errors(errors, response)
                return response

            # Step 4: Transform data with computed columns
            transformed_data = self._transform_data_rows(all_filtered_data, mapping_data, reporting_date, errors)
            if self._handle_errors(errors, response):
                return response

            # Step 5: Create Excel reports and ZIP file
            result_file = self._create_excel_reports(transformed_data, mapping_file, date_str, reporting_date, errors)
            if self._handle_errors(errors, response):
                return response

            # Set the data in the response object
            response.data = result_file
            logger.info(f"Completed processing aging report, returning ZIP file '{result_file.name}'")
            return response

        except Exception as e:
            # Log the error and return error response
            end_time = time.time() - method_start_time
            error_message = f"Error processing file: {str(e)}"
            logger.error(error_message, exc_info=True)
            errors.append(error_message)
            self._handle_errors(errors, response)
            return response
